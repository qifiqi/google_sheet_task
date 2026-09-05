"""回测域 API（自 backtest_training.py / backtest_multi_product.py 归位，URL 不变）。

- bt_api_bp（/backtest-training）与 bmp_api_bp（/backtest-multi-product）两个蓝图
  共存一文件：bt/bmp 的 API 端点同构，页面路由仍留在原页面蓝图；
- import-excel / calculate-ratios 挂保护性限流（06 §3，rate_limit_heavy）。
"""

from __future__ import annotations

import json
import math

from flask import Blueprint, current_app, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.exceptions import BadRequestError, NotFoundError
from app.extensions import limiter
from app.repositories import task_repository, task_result_repository
from app.schemas.backtest import CalculateRatiosSchema
from app.services.backtest_excel_service import BacktestExcelService
from app.services.backtest_multi_product_service import (
    BACKTEST_MULTI_PRODUCT_TASK_TYPE,
    build_multi_product_global_preview_payload,
    normalize_multi_product_config,
)
from app.services.backtest_training_api_service import (
    C3_PARAMETER_FIELDS,
    _build_backtest_result_export_data,
    _build_backtest_result_export_rows,
    _build_c3_summary_rows,
    _build_global_preview_payload,
    _infer_backtest_model_version,
    _load_backtest_task,
    _load_backtest_task_result,
)
from app.services.performance_analysis.historical_metrics import extract_core_metrics
from app.utils.api_response import success
from app.utils.auth import login_required
from app.utils.backtest_report_metadata import get_backtest_model_version, get_price_type
from app.utils.c7_result_normalizer import normalize_c7_result_metrics
from app.utils.request_parsing import parse_body
from app.utils.return_series import parse_return_series_fields
from app.utils.task_types import normalize_task_type


bt_api_bp = Blueprint("backtest_training_api", __name__, url_prefix="/backtest-training")
bmp_api_bp = Blueprint("backtest_multi_product_api", __name__, url_prefix="/backtest-multi-product")


def _rate_limit(config_key, default):
    """限流阈值经 config_manager 运行时可调（零重启）。"""
    from app.services.config_manager import get_config_manager

    return get_config_manager().get_config(config_key, default)


def _user_key():
    from flask import g

    return f"user:{getattr(getattr(g, 'current_user', None), 'id', 'anon')}"


def _sanitize_json_value(value):
    if isinstance(value, float):
        return value if value == value and value not in (float("inf"), float("-inf")) else None
    if isinstance(value, dict):
        return {key: _sanitize_json_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_value(item) for item in value]
    return value


def _parse_json(raw, default):
    if isinstance(raw, (dict, list)):
        return raw
    try:
        return json.loads(raw) if raw else default
    except (TypeError, json.JSONDecodeError):
        return default


# ==================== bt：Excel 导入 / 结果查询 ====================


@bt_api_bp.route("/api/import-excel", methods=["POST"])
@login_required
@limiter.limit(
    lambda: f"{_rate_limit('rate_limit_heavy', 6) or 6}/minute",
    key_func=_user_key,
)
def import_excel():
    excel_file = request.files.get("file")
    if not excel_file or not excel_file.filename:
        raise BadRequestError("请先上传 Excel 文件")

    try:
        data = BacktestExcelService().import_uploaded_excel(excel_file)
    except ValueError as exc:
        raise BadRequestError(str(exc))
    return success(data=_sanitize_json_value(data))


@bt_api_bp.route("/api/task-results/<task_id>", methods=["GET"])
@login_required
def get_task_results_by_task_id(task_id):
    """Return paginated task result summaries for the detail page."""
    _load_backtest_task(task_id)

    page = request.args.get("page", default=1, type=int) or 1
    per_page = request.args.get("per_page", default=10, type=int) or 10
    page = max(page, 1)
    per_page = max(min(per_page, 100), 1)

    page_data = task_result_repository.list_by_task_paginated_raw_parameters(task_id, page, per_page)
    results = [
        {
            **item,
            "parameters": json.loads(item["parameters"]) if item["parameters"] else {},
        }
        for item in page_data["items"]
    ]

    return success(data={
        "task_id": task_id,
        "results": results,
        "pagination": {
            "page": page_data["current_page"],
            "per_page": page_data["per_page"],
            "pages": page_data["pages"],
            "total": page_data["total"],
            "has_prev": page_data["has_prev"],
            "has_next": page_data["has_next"],
            "prev_num": page_data["prev_num"],
            "next_num": page_data["next_num"],
        },
    })


@bt_api_bp.route("/api/task-result/<int:task_result_id>", methods=["GET"])
@login_required
def get_task_result_detail(task_result_id):
    """Return the full task result payload for the result page."""
    task_result, task = _load_backtest_task_result(task_result_id)

    export_data = _build_backtest_result_export_data(task_result, task)

    task_config = task.to_dict().get("config") or {}
    sheet = task_config.get("sheet") if isinstance(task_config.get("sheet"), dict) else {}
    return_series = (
        task_result_repository.get_return_entity(task_result.return_series_id)
        if task_result.return_series_id
        else None
    )
    return success(data={
        "result": _sanitize_json_value(export_data["analyze_result"]),
        "word_report_payload": {
            "report_type": "RPT-S",
            "task_id": task.id,
            "return_series_id": task_result.return_series_id,
            "products": [{
                "stock_code": return_series.stock_code,
                "product_name": return_series.stock_name,
            }] if return_series else [],
            "metadata": {
                "model_version": get_backtest_model_version(
                    sheet.get("title")
                    or task_config.get("title")
                    or task_config.get("spreadsheet_title")
                ),
                "price_type": get_price_type(task_config.get("price_mode") or task_config.get("price_type")),
            },
        } if task_result.return_series_id else None,
    })


@bt_api_bp.route("/api/task-result/<int:task_result_id>/export-preview", methods=["GET"])
@login_required
@limiter.limit(
    lambda: f"{_rate_limit('rate_limit_export', 10) or 10}/minute",
    key_func=_user_key,
)
def get_task_result_export_preview(task_result_id):
    task_result, task = _load_backtest_task_result(task_result_id)

    try:
        export_data = _build_backtest_result_export_data(task_result, task)
        rows = _build_backtest_result_export_rows(export_data)
    except Exception:
        current_app.logger.exception("Failed to build backtest result export preview")
        raise BadRequestError("预览数据生成失败")

    return success(data={
        "filename": export_data["filename"],
        "rows": rows,
    })


@bt_api_bp.route("/api/task-summary/<task_id>", methods=["GET"])
@login_required
def get_task_summary(task_id):
    task = _load_backtest_task(task_id)

    task_config = task.to_dict().get("config") or {}
    model_version = _infer_backtest_model_version(task_config)
    if model_version != "c3":
        raise BadRequestError("当前汇总页仅支持 C3 回测任务")

    rows, parameter_group_count = _build_c3_summary_rows(task_id)

    return success(data={
        "task": {
            "id": task.id,
            "name": task.name,
            "model_version": model_version,
        },
        "parameter_fields": [
            {"key": field_key, "label": field_label}
            for field_key, field_label in C3_PARAMETER_FIELDS
        ],
        "summary": {
            "row_count": len(rows),
            "parameter_group_count": parameter_group_count,
        },
        "rows": _sanitize_json_value(rows),
    })


@bt_api_bp.route("/api/global-preview/<task_id>", methods=["GET"])
@login_required
def get_global_preview(task_id):
    _load_backtest_task(task_id)

    payload = _build_global_preview_payload(task_id)
    if payload is None:
        raise NotFoundError("任务不存在")

    return success(data=_sanitize_json_value(payload))


# ==================== bmp：批量导出构建块（P3 行为级合并的复用件，测试覆盖） ====================


def _build_excel_download_name(task_name, fallback_id: str) -> str:
    safe_name = "".join(char if char not in '\\/:*?"<>|' else "_" for char in str(task_name or "").strip())
    safe_name = safe_name.rstrip(" .")
    return f"{safe_name or fallback_id}.xlsx"


def _build_zip_member_name(task_name: str | None, fallback_id: str, used_names: set[str]) -> str:
    filename = _build_excel_download_name(task_name, fallback_id)
    if filename not in used_names:
        used_names.add(filename)
        return filename

    stem = filename[:-5]
    index = 2
    while True:
        candidate = f"{stem}_{index}.xlsx"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        index += 1


def _parse_excel_percent_text(value: str) -> float | None:
    text = value.strip().replace(",", "").replace("$", "")
    if not text.endswith("%"):
        return None

    sign = 1
    while text.startswith("-"):
        sign *= -1
        text = text[1:]
    try:
        number = sign * float(text[:-1]) / 100
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _format_excel_data_cell(cell):
    if not isinstance(cell.value, str):
        return

    parsed = _parse_excel_percent_text(cell.value)
    if parsed is None:
        return

    cell.value = 0 if parsed == 0 else parsed
    cell.number_format = "0.00%"


def _build_global_preview_workbook(payload: dict[str, object]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "多品全局预览"
    header_fill = PatternFill("solid", fgColor="F7E1A1")
    sub_header_fill = PatternFill("solid", fgColor="FCECC5")
    first_col_fill = PatternFill("solid", fgColor="F7E1A1")
    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)

    products = payload.get("products") or []
    total_columns = max(4, 2 + len(products) * 3 + 2)
    last_column = get_column_letter(total_columns)

    for group in payload.get("groups") or []:
        group_header = ["", ""]
        for product in products:
            group_header.extend([product.get("product_name") or product.get("stock_code") or "产品", "", ""])
        group_header.extend(["", ""])
        sheet.append(group_header[:total_columns])
        group_title_row = sheet.max_row

        current_column = 3
        for _product in products:
            sheet.merge_cells(
                start_row=group_title_row,
                start_column=current_column,
                end_row=group_title_row,
                end_column=current_column + 2,
            )
            current_column += 3

        header = ["指标类型", "指标"]
        for product in products:
            ratio = product.get("ratio")
            header.extend(["指数", "模型结果", f"模型结果（{ratio}%）"])
        header.extend(["比例计算-指数", "比例计算-结果"])
        sheet.append(header)
        for row in group.get("rows") or []:
            values = [row.get("category") or "", row.get("metric") or ""]
            for product_value in row.get("product_values") or []:
                values.extend([
                    product_value.get("index_value") or "-",
                    product_value.get("result_value") or "-",
                    product_value.get("weighted_result_value") or "-",
                ])
            values.extend([
                row.get("weighted_index_value") or "-",
                row.get("weighted_result_value") or "-",
            ])
            sheet.append(values)
        sheet.append([""] * total_columns)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin_border
            cell.font = body_font
            if cell.row == 1:
                cell.fill = sub_header_fill
                cell.font = header_font
            if cell.value in {
                "指标类型",
                "指标",
                "指数",
                "模型结果",
                "比例计算-指数",
                "比例计算-结果",
            } or str(cell.value or "").startswith("模型结果（"):
                cell.font = header_font
                cell.fill = sub_header_fill
            if cell.column == 1:
                cell.fill = first_col_fill
                if cell.row <= 2:
                    cell.font = header_font
            if cell.row >= 3 and cell.column >= 3:
                _format_excel_data_cell(cell)
    for column_index in range(1, total_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18 if column_index > 2 else 16
    sheet.freeze_panes = "A3"
    if sheet.max_row >= 2:
        sheet.auto_filter.ref = f"A2:{last_column}{sheet.max_row}"
    return workbook


# ==================== bmp：Excel 导入 / 结果查询 / 全局预览 ====================


def _infer_product_export_model_name(product):
    if not isinstance(product, dict):
        return "C3"

    sheet = product.get("sheet") or {}
    for source in (product.get("model_name"), sheet.get("title"), product.get("model_version")):
        title = str(source or "").upper()
        for model_name in ("C7", "C5", "C4", "C3"):
            if model_name in title:
                return model_name
    return "C3"


def _build_word_report_payload(task: dict, task_result) -> dict | None:
    """按当前结果的参数方案构造多品 Word 报告请求。"""
    try:
        config = normalize_multi_product_config(task.get("config") or {})
    except ValueError:
        return None
    selected_parameters = _parse_json(task_result.parameters, {})
    group_index = str(selected_parameters.get("parameter_group_index") or 0)
    return {
        "report_type": "RPT-M",
        "task_id": task["id"],
        "group_key": group_index,
        # normalize_multi_product_config 已把历史布尔配置归一为 weighting_mode。
        "weighting_mode": config.get("weighting_mode") or "daily_compound",
        "ratios": [
            {"product_index": product["product_index"], "ratio": product["ratio"]}
            for product in config["products"]
        ],
    }


@bmp_api_bp.route("/api/import-excel", methods=["POST"])
@login_required
@limiter.limit(
    lambda: f"{_rate_limit('rate_limit_heavy', 6) or 6}/minute",
    key_func=_user_key,
)
def bmp_import_excel():
    excel_file = request.files.get("file")
    if not excel_file or not excel_file.filename:
        raise BadRequestError("请先上传 Excel 文件")
    try:
        data = BacktestExcelService().import_uploaded_excel(excel_file)
    except ValueError as exc:
        raise BadRequestError(str(exc))
    return success(data=_sanitize_json_value(data))


@bmp_api_bp.route("/api/task-results/<task_id>", methods=["GET"])
@login_required
def bmp_get_task_results_by_task_id(task_id):
    _load_multi_product_task_or_none(task_id)

    page = max(request.args.get("page", default=1, type=int) or 1, 1)
    per_page = max(min(request.args.get("per_page", default=10, type=int) or 10, 100), 1)
    page_data = task_result_repository.list_by_task_paginated_raw_parameters(task_id, page, per_page)
    results = [
        {**item, "parameters": _parse_json(item["parameters"], {})}
        for item in page_data["items"]
    ]
    return success(data={
        "task_id": task_id,
        "results": results,
        "pagination": {
            "page": page_data["current_page"],
            "per_page": page_data["per_page"],
            "pages": page_data["pages"],
            "total": page_data["total"],
            "has_prev": page_data["has_prev"],
            "has_next": page_data["has_next"],
            "prev_num": page_data["prev_num"],
            "next_num": page_data["next_num"],
        },
    })


@bmp_api_bp.route("/api/task-result/<int:task_result_id>", methods=["GET"])
@login_required
def bmp_get_task_result_detail(task_result_id):
    task_result = task_result_repository.get_entity(task_result_id)
    if not task_result:
        raise NotFoundError("任务结果不存在")
    task = _load_multi_product_task_or_none(task_result.task_id)

    payload = _parse_json(task_result.result, {})
    if isinstance(payload, dict) and payload:
        prioritized_keys = ("metrics_payload", "calculate_metrics", "weighted_calculate_metrics", "analyze_result")
        value = next(
            (
                item
                for item in payload.values()
                if isinstance(item, dict) and any(key in item for key in prioritized_keys)
            ),
            next((item for item in payload.values() if isinstance(item, dict)), {}),
        )
    else:
        value = {}
    calculate_metrics = extract_core_metrics(value)
    sheet_result = {
        key: item
        for key, item in value.items()
        if key not in {"metrics_payload", "calculate_metrics", "analyze_result"}
    } if isinstance(value, dict) else {}

    daily_returns = {}
    if task_result.return_series_id:
        return_series = task_result_repository.get_return_entity(task_result.return_series_id)
        if return_series:
            rows = parse_return_series_fields(return_series)
            daily_returns = {
                "dates": [row["date"] for row in rows],
                "index_returns": [row.get("index_return") for row in rows],
                "start_returns": [row.get("start_return") for row in rows],
            }

    task_config = _parse_json(task.get("config"), {})
    products = task_config.get("products") if isinstance(task_config, dict) else []
    parameters = _parse_json(task_result.parameters, {})
    product_index = parameters.get("product_index") if isinstance(parameters, dict) else None
    product = products[product_index] if isinstance(product_index, int) and 0 <= product_index < len(products) else {}
    model_name = _infer_product_export_model_name(product)
    if model_name == "C7":
        sheet_result = normalize_c7_result_metrics(sheet_result)

    return success(data={
        "result": _sanitize_json_value({
            **(calculate_metrics if isinstance(calculate_metrics, dict) else {}),
            "sheet_result": sheet_result,
            "daily_returns": daily_returns,
            "model_name": model_name,
        }),
        "word_report_payload": _build_word_report_payload(task, task_result),
    })


@bmp_api_bp.route("/api/global-preview/<task_id>", methods=["GET"])
@login_required
def bmp_get_global_preview(task_id):
    _load_multi_product_task_or_none(task_id)
    payload = build_multi_product_global_preview_payload(task_id)
    if payload is None:
        raise NotFoundError("任务不存在")
    return success(data=_sanitize_json_value(payload))


@bmp_api_bp.route("/api/global-preview/<task_id>/calculate-ratios", methods=["POST"])
@login_required
@limiter.limit(
    lambda: f"{_rate_limit('rate_limit_heavy', 6) or 6}/minute",
    key_func=_user_key,
)
def bmp_calculate_ratios(task_id):
    _load_multi_product_task_or_none(task_id)
    data = parse_body(CalculateRatiosSchema)
    ratios = data.ratios
    try:
        payload = build_multi_product_global_preview_payload(task_id, ratios_override=ratios)
    except ValueError as exc:
        raise BadRequestError(str(exc))
    if payload is None:
        raise NotFoundError("任务不存在")
    return success(data=_sanitize_json_value(payload))


@bmp_api_bp.route("/api/global-preview/<task_id>/ratios", methods=["PUT"])
@login_required
def bmp_update_ratios(task_id):
    task = _load_multi_product_task_or_none(task_id)
    data = request.get_json() or {}
    ratios = data.get("ratios")
    if not isinstance(ratios, list):
        raise BadRequestError("ratios 必须是数组")

    config = normalize_multi_product_config(task.get("config") or {})
    products = config["products"]
    if len(ratios) != len(products):
        raise BadRequestError("比例数量与产品数量不一致")
    for product, ratio in zip(products, ratios):
        product["ratio"] = str(ratio.get("ratio") if isinstance(ratio, dict) else ratio).strip()
    try:
        config = normalize_multi_product_config({**config, "products": products})
    except ValueError as exc:
        raise BadRequestError(str(exc))

    task_repository.update_fields(task_id, config=json.dumps(config, ensure_ascii=False))
    payload = build_multi_product_global_preview_payload(task_id)
    return success(
        data=_sanitize_json_value(payload or {}),
        message="比例已保存",
    )


def _load_multi_product_task_or_none(task_id: str):
    """加载多品回测任务 dict；不存在返回 None，类型不符抛 BadRequestError。"""
    task = task_repository.get(task_id)
    if not task:
        return None
    if normalize_task_type(task["task_type"]) != BACKTEST_MULTI_PRODUCT_TASK_TYPE:
        raise BadRequestError("当前接口仅支持多品数据回测任务")
    return task
