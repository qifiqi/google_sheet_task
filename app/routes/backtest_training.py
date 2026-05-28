"""Backtest training page routes."""

import json
import math
import re
from collections import OrderedDict
from decimal import Decimal, InvalidOperation
from io import BytesIO

from flask import Blueprint, current_app, jsonify, render_template, request, send_file, g
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import load_only

from app.extensions import db
from app.models import Task, TaskResult
from app.services.backtest_excel_service import BacktestExcelService
from app.services.stock_metadata_service import bulk_upsert_stock_metadata
from app.services.xpl_service import xpl_analyzer
from app.utils.dfcf_api import DFCJStockApi
from app.utils.auth import login_required, permission_required
from app.utils.task_authorization import authorize_task_type_action, normalize_task_type

bp = Blueprint("backtest_training", __name__, url_prefix="/backtest-training")
legacy_bp = Blueprint("backtest_training_legacy", __name__, url_prefix="/backtest")

C3_PARAMETER_FIELDS = [
    ("commission", "Commission"),
    ("xm", "X Multiplier"),
    ("dbbh1", "单边保护1"),
    ("dbbh2", "单边保护2"),
    ("zlxc", "中立限仓"),
    ("zsgz", "指数跟踪"),
    ("ywf1", "一窝蜂 smoothing"),
    ("ywf2", "一窝蜂 bordering"),
]

TASK_ACTION_LABELS = {
    "view": "查看",
}

SCIENTIFIC_NOTATION_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)[eE][+-]?\d+$")

SUMMARY_METRIC_CELL_MAP = {
    "C3": {
        "index_return": "I18",
        "return": "I15",
        "index_max_drawdown": "I20",
        "max_drawdown": "I17",
    },
    "C5": {
        "index_return": "D5",
        "return": "D2",
        "index_max_drawdown": "D7",
        "max_drawdown": "D4",
    },
}
SUMMARY_METRIC_CELL_MAP["C4"] = SUMMARY_METRIC_CELL_MAP["C5"]

SUMMARY_ROW_LABELS = [
    ("index_return", "指数回报"),
    ("return", "模型回报"),
    ("excess_return", "超额回报"),
    ("index_max_drawdown", "指数回撤"),
    ("max_drawdown", "模型回撤"),
    ("excess_drawdown", "超额回撤"),
]


def _normalize_scientific_text(text: str) -> str:
    if not SCIENTIFIC_NOTATION_RE.fullmatch(text):
        return text

    try:
        number = Decimal(text)
    except InvalidOperation:
        return text

    if not number.is_finite():
        return text

    normalized = format(number.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    if normalized in {"-0", "+0"}:
        return "0"
    return normalized


def _task_permission_denied(action: str, task_type: str | None, decision: dict, task_id: str | None = None, result_id: int | None = None):
    action_label = TASK_ACTION_LABELS.get(action, action)
    normalized_type = decision.get("task_type") or str(task_type or "unknown")
    missing_permissions = decision.get("missing_permissions") or []
    missing_text = "、".join(missing_permissions) if missing_permissions else "未知"
    message = f"权限不足，无法{action_label}{normalized_type}任务；当前缺少: {missing_text}"

    return jsonify({
        "status": "error",
        "message": message,
        "action": action,
        "task_type": normalized_type,
        "task_id": task_id,
        "result_id": result_id,
        "required_permissions": decision.get("required_permissions") or [],
        "missing_permissions": missing_permissions,
    }), 403


def _load_backtest_task_or_response(task_id: str, action: str = "view", result_id: int | None = None):
    task = db.session.get(Task, task_id)
    if not task:
        return None, (jsonify({
            "status": "error",
            "message": "任务不存在",
        }), 404)

    decision = authorize_task_type_action(getattr(g, "current_user", None), action, task.task_type)
    if not decision["allowed"]:
        return None, _task_permission_denied(action, task.task_type, decision, task_id=task_id, result_id=result_id)

    if normalize_task_type(task.task_type) != "backtest_training":
        return None, (jsonify({
            "status": "error",
            "message": "当前接口仅支持回测任务",
            "task_id": task_id,
            "task_type": task.task_type,
        }), 400)

    return task, None


def _sanitize_json_value(value):
    """Convert NaN/Infinity values into JSON-safe nulls."""
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {key: _sanitize_json_value(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_value(item) for item in value]
    return value


def _strip_html_tags(value):
    return re.sub(r"<[^>]+>", "", str(value or "")).strip()


def _infer_backtest_model_version(config):
    if not isinstance(config, dict):
        return "c3"

    sheet = config.get("sheet") or {}
    title = str(sheet.get("title") or config.get("title") or "").upper()
    if "C5" in title or "C4" in title:
        return "c5"

    parameters = config.get("parameters") or []
    first_row = parameters[0] if parameters and isinstance(parameters[0], list) else []
    if len(first_row) == 2:
        return "c5"
    return "c3"


def _parse_percent_like_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value if math.isfinite(value) else None

    raw = str(value).strip()
    if not raw or raw == "-":
        return None

    normalized = raw.replace(",", "").replace("$", "")
    try:
        if normalized.endswith("%"):
            return float(normalized[:-1]) / 100
        return float(normalized)
    except (TypeError, ValueError):
        return raw


def _extract_task_result_payload(task_result):
    try:
        result_payload = json.loads(task_result.result) if task_result.result else {}
    except (TypeError, json.JSONDecodeError):
        result_payload = {}

    if isinstance(result_payload, dict) and result_payload:
        value = next(
            (
                item
                for item in result_payload.values()
                if isinstance(item, dict) and "calculate_metrics" in item
            ),
            next((item for item in result_payload.values() if isinstance(item, dict)), {}),
        )
    else:
        value = {}
    if not isinstance(value, dict):
        return {}, {}

    calculate_metrics = value.get("calculate_metrics")
    sheet_result = {
        key: item
        for key, item in value.items()
        if key != "calculate_metrics"
    }
    return (
        calculate_metrics if isinstance(calculate_metrics, dict) else {},
        sheet_result,
    )


def _extract_year_drawdown_map(section):
    if not isinstance(section, dict):
        return {}
    return {
        str(item.get("year")): item
        for item in section.get("year_maximum_drawdown", [])
        if isinstance(item, dict) and item.get("year") not in (None, "", "all")
    }


def _extract_year_sharpe_map(section):
    if not isinstance(section, dict):
        return {}

    year_map = {}
    for key, value in section.items():
        if not isinstance(value, dict):
            continue
        match = re.match(r"^year_\d+_(\d{4})$", str(key))
        if match:
            year_map[match.group(1)] = value
    return year_map


def _extract_display_year(source_window):
    raw = str(source_window or "").strip()
    if not raw:
        return ""

    if re.fullmatch(r"\d{4}-\d{4}", raw):
        end_year, start_year = raw.split("-", 1)
        return start_year or end_year

    return raw


def _build_c3_summary_rows(task_id):
    task_results = (
        TaskResult.query
        .options(
            load_only(
                TaskResult.id,
                TaskResult.task_id,
                TaskResult.step_index,
                TaskResult.parameters,
                TaskResult.result,
                TaskResult.success,
                TaskResult.timestamp,
            )
        )
        .filter_by(task_id=task_id, success=True)
        .order_by(TaskResult.step_index.asc(), TaskResult.timestamp.asc(), TaskResult.id.asc())
        .all()
    )

    rows = []

    for task_result in task_results:
        try:
            parameters = json.loads(task_result.parameters) if task_result.parameters else {}
        except (TypeError, json.JSONDecodeError):
            parameters = {}

        if not isinstance(parameters, dict):
            continue

        parameter_values = parameters.get("parameter")
        if not isinstance(parameter_values, list) or not parameter_values:
            continue

        parameter_map = {}
        for index, (field_key, _field_label) in enumerate(C3_PARAMETER_FIELDS):
            parameter_map[field_key] = parameter_values[index] if index < len(parameter_values) else None

        calculate_metrics, sheet_result = _extract_task_result_payload(task_result)
        def _safe_all_entry(items, key_name="year"):
            if not isinstance(items, list):
                return {}
            for item in items:
                if isinstance(item, dict) and str(item.get(key_name)) == "all":
                    return item
            return {}

        index_sharpe_all = (
            (calculate_metrics.get("index_sharpe_ratios") or {}).get("all")
            if isinstance(calculate_metrics.get("index_sharpe_ratios"), dict)
            else {}
        ) or {}
        start_sharpe_all = (
            (calculate_metrics.get("start_sharpe_ratios") or {}).get("all")
            if isinstance(calculate_metrics.get("start_sharpe_ratios"), dict)
            else {}
        ) or {}
        excess_returns = calculate_metrics.get("excess_returns") or []
        all_excess = next(
            (
                item for item in excess_returns
                if isinstance(item, dict) and str(item.get("year")) == "all"
            ),
            {}
        )
        index_profit_monthly_all = _safe_all_entry(
            calculate_metrics.get("index_profit_monthly")
        )
        start_profit_monthly_all = _safe_all_entry(
            calculate_metrics.get("start_profit_monthly")
        )
        index_kama_all = _safe_all_entry(calculate_metrics.get("index_kama_ratio"))
        start_kama_all = _safe_all_entry(calculate_metrics.get("start_kama_ratio"))
        index_sotino_all = _safe_all_entry(
            calculate_metrics.get("index_sotino_ratio")
        )
        start_sotino_all = _safe_all_entry(
            calculate_metrics.get("start_sotino_ratio")
        )
        monthly_excess_percentage_all = _safe_all_entry(
            calculate_metrics.get("monthly_excess_return_percentage")
        )
        monthly_excess_returns = calculate_metrics.get("monthly_excess_returns") or []
        monthly_excess_values = [
            item.get("monthly_excess_return_diff")
            for item in monthly_excess_returns
            if isinstance(item, dict)
            and item.get("monthly_excess_return_diff") is not None
        ]
        avg_monthly_excess_return = (
            sum(monthly_excess_values) / len(monthly_excess_values)
            if monthly_excess_values
            else None
        )

        source_window = str(parameters.get("year") or parameters.get("Kline_key") or "")
        year_label = _extract_display_year(source_window)
        parameter_signature = json.dumps(parameter_values, ensure_ascii=False)

        strategy_return = _parse_percent_like_value(sheet_result.get("I15"))
        index_return = _parse_percent_like_value(sheet_result.get("I18"))
        beats_index = None
        if isinstance(strategy_return, (int, float)) and isinstance(index_return, (int, float)):
            beats_index = strategy_return - index_return
        else:
            beats_index = _parse_percent_like_value(all_excess.get("annualized_return_diff"))

        strategy_max_drawdown = _parse_percent_like_value(sheet_result.get("I17"))
        index_max_drawdown = _parse_percent_like_value(sheet_result.get("I20"))
        drawdown_beats = None
        if isinstance(strategy_max_drawdown, (int, float)) and isinstance(index_max_drawdown, (int, float)):
            drawdown_beats = strategy_max_drawdown - index_max_drawdown

        rows.append({
            **parameter_map,
            "year": year_label,
            "strategy_return": strategy_return,
            "strategy_annualized": _parse_percent_like_value(sheet_result.get("I16")),
            "index_return": index_return,
            "index_annualized": _parse_percent_like_value(sheet_result.get("I19")),
            "beats_index": beats_index,
            "strategy_max_drawdown": strategy_max_drawdown,
            "index_max_drawdown": index_max_drawdown,
            "drawdown_beats": drawdown_beats,
            "fee_total": _parse_percent_like_value(sheet_result.get("I21")),
            "fee_annualized": _parse_percent_like_value(sheet_result.get("I22")),
            "year_rate": _parse_percent_like_value(sheet_result.get("I23")),
            "index_monthly_sharpe": _parse_percent_like_value(index_sharpe_all.get("sharpe_ratio")),
            "strategy_monthly_sharpe": _parse_percent_like_value(start_sharpe_all.get("sharpe_ratio")),
            "index_avg_monthly_return": _parse_percent_like_value(
                index_sharpe_all.get("avg_monthly_return")
            ),
            "strategy_avg_monthly_return": _parse_percent_like_value(
                start_sharpe_all.get("avg_monthly_return")
            ),
            "index_monthly_return_volatility": _parse_percent_like_value(
                calculate_metrics.get("index_monthly_return_volatility")
            ),
            "strategy_monthly_return_volatility": _parse_percent_like_value(
                calculate_metrics.get("start_monthly_return_volatility")
            ),
            "excess_annualized_return": _parse_percent_like_value(
                all_excess.get("annualized_return_diff")
            ),
            "outperform_year": _parse_percent_like_value(
                calculate_metrics.get("outperform_year")
            ),
            "monthly_excess_return_percentage": _parse_percent_like_value(
                monthly_excess_percentage_all.get("excess_return")
            ),
            "avg_monthly_excess_return": _parse_percent_like_value(
                avg_monthly_excess_return
            ),
            "monthly_excess_volatility": _parse_percent_like_value(
                calculate_metrics.get("monthly_excess_volatility")
            ),
            "index_profit_annual": _parse_percent_like_value(
                calculate_metrics.get("index_profit_annual")
            ),
            "strategy_profit_annual": _parse_percent_like_value(
                calculate_metrics.get("start_profit_annual")
            ),
            "index_profit_monthly_percentage": _parse_percent_like_value(
                index_profit_monthly_all.get("profit_monthly_percentage")
            ),
            "strategy_profit_monthly_percentage": _parse_percent_like_value(
                start_profit_monthly_all.get("profit_monthly_percentage")
            ),
            "index_kama_ratio": _parse_percent_like_value(
                index_kama_all.get("kama_ratio")
            ),
            "strategy_kama_ratio": _parse_percent_like_value(
                start_kama_all.get("kama_ratio")
            ),
            "index_sotino_ratio": _parse_percent_like_value(
                index_sotino_all.get("sotino_ratio")
            ),
            "strategy_sotino_ratio": _parse_percent_like_value(
                start_sotino_all.get("sotino_ratio")
            ),
            "excess_sharp": _parse_percent_like_value(
                calculate_metrics.get("excess_sharp")
            ),
            "excess_of_promissory_note": _parse_percent_like_value(
                calculate_metrics.get("excess_of_promissory_note")
            ),
            "excess_drawdown_winning_rate": _parse_percent_like_value(
                calculate_metrics.get("excess_drawdown_winning_rate")
            ),
            "index_maximum_number_of_backtest_repair_days": _parse_percent_like_value(
                calculate_metrics.get("index_maximum_number_of_backtest_repair_days")
            ),
            "strategy_maximum_number_of_backtest_repair_days": _parse_percent_like_value(
                calculate_metrics.get("start_maximum_number_of_backtest_repair_days")
            ),
            "excess_maximum_number_of_backtest_repair_days": _parse_percent_like_value(
                calculate_metrics.get("excess_maximum_number_of_backtest_repair_days")
            ),
            "date_range": all_excess.get("start_end_date"),
            "source_window": source_window,
            "task_result_id": task_result.id,
            "step_index": task_result.step_index,
            "timestamp": task_result.timestamp.isoformat() if task_result.timestamp else None,
            "parameter_signature": parameter_signature,
        })

    rows.sort(
        key=lambda item: (
            item.get("parameter_signature") or "",
            -(int(item.get("year")) if str(item.get("year", "")).isdigit() else -9999),
            item.get("step_index") or 0,
        )
    )

    parameter_group_count = len({row["parameter_signature"] for row in rows})
    for row in rows:
        row.pop("parameter_signature", None)

    return rows, parameter_group_count


@bp.route("/create")
def create_page():
    return render_template("backtest_training/create.html")


@bp.route("/list")
def list_page():
    return render_template("backtest_training/list.html")


@bp.route("/detail/<task_id>")
def detail_page(task_id):
    return render_template("backtest_training/detail.html", task_id=task_id)


@bp.route("/global-preview/<task_id>")
def global_preview_page(task_id):
    return render_template("backtest_training/global_preview.html", task_id=task_id)


@bp.route("/result/<int:result_id>")
def result_page(result_id):
    task_result = TaskResult.query.get(result_id)
    task_id = ""
    if task_result and task_result.task and normalize_task_type(task_result.task.task_type) == "backtest_training":
        task_id = task_result.task_id
    return render_template("backtest_training/result.html", result_id=result_id, task_id=task_id)


legacy_bp.add_url_rule("/create", view_func=create_page)
legacy_bp.add_url_rule("/list", view_func=list_page)
legacy_bp.add_url_rule("/detail/<task_id>", view_func=detail_page)
legacy_bp.add_url_rule("/global-preview/<task_id>", view_func=global_preview_page)
legacy_bp.add_url_rule("/result/<int:result_id>", view_func=result_page)


@bp.route("/api/import-excel", methods=["POST"])
@login_required
@permission_required('backtest:create')
def import_excel():
    excel_file = request.files.get("file")
    if not excel_file or not excel_file.filename:
        return jsonify({
            "status": "error",
            "message": "请先上传 Excel 文件",
        }), 400

    try:
        data = BacktestExcelService().import_uploaded_excel(excel_file)
        return jsonify({
            "status": "success",
            **_sanitize_json_value(data),
        })
    except ValueError as exc:
        return jsonify({
            "status": "error",
            "message": str(exc),
        }), 400
    except Exception as exc:
        current_app.logger.exception("Failed to import backtest Excel")
        return jsonify({
            "status": "error",
            "message": f"Excel 解析失败：{str(exc)}",
        }), 500


@bp.route("/api/search-stocks", methods=["GET"])
@login_required
@permission_required('backtest:view')
def search_stocks():
    keyword = (request.args.get("q") or "").strip()
    page_size = request.args.get("page_size", default=10, type=int) or 10
    page_size = max(1, min(page_size, 20))

    if len(keyword) < 1:
        return jsonify({
            "status": "success",
            "keyword": keyword,
            "results": [],
        })

    raw_results = DFCJStockApi().get_search_list_by_stock_code(keyword, page_size=page_size)
    if isinstance(raw_results, dict) and raw_results.get("error"):
        return jsonify({
            "status": "error",
            "message": raw_results.get("error") or "股票搜索失败",
        }), 502

    normalized_results = []
    for item in raw_results or []:
        if item.get("status") not in (10, "10", None):
            continue
        code = _strip_html_tags(item.get("code"))
        short_name = _strip_html_tags(item.get("shortName"))
        security_type_name = _strip_html_tags(item.get("securityTypeName"))
        market = item.get("market")
        if not code:
            continue
        normalized_results.append({
            "source": item.get("source"),
            "code": code,
            "name": short_name,
            "security_type_name": security_type_name,
            "market": market,
            "is_exact_match": bool(item.get("isExactMatch")),
            "label": " · ".join(part for part in [code, short_name, security_type_name] if part),
            "status": item.get("status"),
            "inner_code": item.get("innerCode"),
            "pinyin": item.get("pinyin"),
            "security_type": item.get("securityType"),
            "small_type": item.get("smallType"),
            "flag": item.get("flag"),
            "ext_small_type": item.get("extSmallType"),
            "quote_id": item.get("quoteId"),
            "market_type": item.get("marketType"),
            "unified_code": item.get("unifiedCode"),
            "jys": item.get("jys"),
            "classify": item.get("classify"),
        })

    bulk_upsert_stock_metadata([
        {
            "stock_code": item.get("code"),
            "stock_name": item.get("name"),
            "market_type": item.get("market_type") or item.get("market"),
            "exchange_market": item.get("market"),
            "security_type_name": item.get("security_type_name"),
            "source": item.get("source"),
            "raw": item,
        }
        for item in normalized_results
    ])
    db.session.commit()

    return jsonify({
        "status": "success",
        "keyword": keyword,
        "results": normalized_results,
    })


@bp.route("/api/task-results/<task_id>", methods=["GET"])
@login_required
@permission_required('backtest:view')
def get_task_results_by_task_id(task_id):
    """Return paginated task result summaries for the detail page."""
    _, error_response = _load_backtest_task_or_response(task_id, action="view")
    if error_response:
        return error_response

    page = request.args.get("page", default=1, type=int) or 1
    per_page = request.args.get("per_page", default=10, type=int) or 10
    page = max(page, 1)
    per_page = max(min(per_page, 100), 1)

    pagination = (
        TaskResult.query
        .options(
            load_only(
                TaskResult.id,
                TaskResult.task_id,
                TaskResult.step_index,
                TaskResult.parameters,
                TaskResult.success,
                TaskResult.error_message,
                TaskResult.timestamp,
            )
        )
        .filter_by(task_id=task_id)
        .order_by(TaskResult.step_index.asc(), TaskResult.timestamp.asc(), TaskResult.id.asc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    results = [
        {
            "id": task_result.id,
            "task_id": task_result.task_id,
            "step_index": task_result.step_index,
            "parameters": json.loads(task_result.parameters) if task_result.parameters else {},
            "success": task_result.success,
            "error_message": task_result.error_message,
            "timestamp": task_result.timestamp.isoformat() if task_result.timestamp else None,
        }
        for task_result in pagination.items
    ]

    return jsonify({
        "status": "success",
        "task_id": task_id,
        "results": results,
        "pagination": {
            "page": pagination.page,
            "per_page": per_page,
            "pages": pagination.pages,
            "total": pagination.total,
            "has_prev": pagination.has_prev,
            "has_next": pagination.has_next,
            "prev_num": pagination.prev_num,
            "next_num": pagination.next_num,
        },
    })


@bp.route("/api/task-result/<int:task_result_id>", methods=["GET"])
@login_required
@permission_required('backtest:view')
def get_task_result_detail(task_result_id):
    """Return the full task result payload for the result page."""
    task_result = (
        TaskResult.query
        .options(
            load_only(
                TaskResult.task_id,
                TaskResult.result
            )
        )
        .filter(TaskResult.id == task_result_id)
        .first()
    )
    if not task_result:
        return jsonify({
            "status": "error",
            "message": "任务结果不存在",
        }), 404

    _, error_response = _load_backtest_task_or_response(task_result.task_id, action="view", result_id=task_result_id)
    if error_response:
        return error_response

    try:
        result_payload = json.loads(task_result.result) if task_result.result else {}
    except (TypeError, json.JSONDecodeError):
        result_payload = {}
    if isinstance(result_payload, dict) and result_payload:
        val = next(
            (
                item
                for item in result_payload.values()
                if isinstance(item, dict) and "calculate_metrics" in item
            ),
            next((item for item in result_payload.values() if isinstance(item, dict)), {}),
        )
    else:
        val = {}
    calculate_metrics = val.get("calculate_metrics") if isinstance(val, dict) else {}
    sheet_result = {
        key: item
        for key, item in val.items()
        if key != "calculate_metrics"
    } if isinstance(val, dict) else {}

    return jsonify({
        "status": "success",
        "result": _sanitize_json_value({
            **(calculate_metrics if isinstance(calculate_metrics, dict) else {}),
            "sheet_result": sheet_result,
        }),
    })

@bp.route("/api/task-summary/<task_id>", methods=["GET"])
@login_required
@permission_required('backtest:view')
def get_task_summary(task_id):
    task, error_response = _load_backtest_task_or_response(task_id, action="view")
    if error_response:
        return error_response

    task_config = task.to_dict().get("config") or {}
    model_version = _infer_backtest_model_version(task_config)
    if model_version != "c3":
        return jsonify({
            "status": "error",
            "message": "当前汇总页仅支持 C3 回测任务",
        }), 400

    rows, parameter_group_count = _build_c3_summary_rows(task_id)

    return jsonify({
        "status": "success",
        "task": {
            "id": task.id,
            "name": task.name,
            "model_version": model_version,
        },
        "parameter_fields": [
            {"key": field_key, "label": field_label}
            for field_key, field_label in C3_PARAMETER_FIELDS
        ],
        "summary": {
            "row_count": len(rows),
            "parameter_group_count": parameter_group_count,
        },
        "rows": _sanitize_json_value(rows),
    })


def _build_parameter_header(parameters):
    parameter_values = parameters.get("parameter")
    if not isinstance(parameter_values, list) or not parameter_values:
        return "未命名参数"

    if len(parameter_values) == 2:
        labels = ["xm", "ml"]
    else:
        labels = [f"参数{i + 1}" for i in range(len(parameter_values))]

    parts = []
    for _, value in zip(labels, parameter_values):
        parts.append(f"{value}")
    return " / ".join(parts)


def _extract_result_core(task_result):
    payload = task_result.to_dict().get("result") or {}
    if not isinstance(payload, dict) or not payload:
        return {}
    first_value = next(iter(payload.values()), {})
    return first_value if isinstance(first_value, dict) else {}


def _detect_model_name(task_name, parameters):
    upper_name = (task_name or "").upper()
    for model_name in ("C5", "C4", "C3"):
        if model_name in upper_name:
            return model_name

    parameter_values = parameters.get("parameter")
    if isinstance(parameter_values, list) and len(parameter_values) == 2:
        return "C5"
    return "C3"


def _extract_raw_sheet_metrics(result_core):
    if not isinstance(result_core, dict):
        return {}
    return {
        str(key): value
        for key, value in result_core.items()
        if key != "calculate_metrics"
    }


def _normalize_summary_numeric_value(value):
    parsed = _parse_percent_like_value(value)
    if isinstance(parsed, (int, float)):
        return parsed if math.isfinite(parsed) else None
    return None


def _format_summary_value(value):
    parsed = _parse_percent_like_value(value)
    if isinstance(parsed, (int, float)):
        if not math.isfinite(parsed):
            return ""
        return f"{parsed:.2%}"
    if parsed is None:
        return ""
    return _normalize_scientific_text(str(parsed).strip())


def _get_summary_raw_metric(column, metric_key):
    model_name = str(column.get("model_name") or "C3").upper()
    cell_map = SUMMARY_METRIC_CELL_MAP.get(model_name, SUMMARY_METRIC_CELL_MAP["C3"])
    cell_key = cell_map.get(metric_key)
    raw_metrics = column.get("raw_metrics") or {}
    return raw_metrics.get(cell_key) if cell_key else None


def _get_summary_derived_value(column, metric_key):
    if metric_key == "excess_return":
        left = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "return")
        )
        right = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "index_return")
        )
        if left is None or right is None:
            return ""
        return f"{left - right:.2%}"

    if metric_key == "excess_drawdown":
        left = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "max_drawdown")
        )
        right = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "index_max_drawdown")
        )
        if left is None or right is None:
            return ""
        return f"{left - right:.2%}"

    return _format_summary_value(_get_summary_raw_metric(column, metric_key))


def _extract_summary_rows(calculate_metrics, model_name):
    if not isinstance(calculate_metrics, dict) or not calculate_metrics:
        return "", []

    def _normalize_metric_label(label):
        text = str(label or "").strip()
        text = text.replace("（", "(").replace("）", ")")
        metric_aliases = {
            "跑赢年份(百分比)": "跑赢年份(百分比)",
            "跑赢年份(百分比 )": "跑赢年份(百分比)",
            "超额最大修复天数": "超额最大修复天数",
            "最大修复天数": "最大修复天数",
            "所提诺比率": "所提诺比率",
            "超额所提诺比率": "超额所提诺比率",
        }
        return metric_aliases.get(text, text)

    def _normalize_display_value(value):
        text = str(value or "").strip()
        if not text:
            return ""
        while text.startswith("--"):
            text = text[1:]
        return _normalize_scientific_text(text)

    def _fmt_percent(value):
        if value is None or not math.isfinite(value):
            return ""
        return f"{value:.2%}"

    def _fmt_number(value):
        if value is None or not math.isfinite(value):
            return ""
        return f"{value:.2f}".rstrip("0").rstrip(".")

    def _safe_all_entry(items, key_name):
        if not isinstance(items, list):
            return {}
        for item in items:
            if isinstance(item, dict) and str(item.get(key_name)) == "all":
                return item
        return {}

    def _build_fallback_rows():
        excess_all = _safe_all_entry(calculate_metrics.get("excess_returns"), "year")
        index_profit_monthly_all = _safe_all_entry(calculate_metrics.get("index_profit_monthly"), "year")
        start_profit_monthly_all = _safe_all_entry(calculate_metrics.get("start_profit_monthly"), "year")
        index_kama_all = _safe_all_entry(calculate_metrics.get("index_kama_ratio"), "year")
        start_kama_all = _safe_all_entry(calculate_metrics.get("start_kama_ratio"), "year")
        index_sotino_all = _safe_all_entry(calculate_metrics.get("index_sotino_ratio"), "year")
        start_sotino_all = _safe_all_entry(calculate_metrics.get("start_sotino_ratio"), "year")
        monthly_excess_percentage_all = _safe_all_entry(
            calculate_metrics.get("monthly_excess_return_percentage"), "year"
        )
        index_sharpe_all = (calculate_metrics.get("index_sharpe_ratios") or {}).get("all") or {}
        start_sharpe_all = (calculate_metrics.get("start_sharpe_ratios") or {}).get("all") or {}

        monthly_excess_returns = calculate_metrics.get("monthly_excess_returns") or []
        valid_excess_months = [
            item.get("monthly_excess_return_diff")
        for item in monthly_excess_returns
            if isinstance(item, dict) and item.get("monthly_excess_return_diff") is not None
        ]
        avg_monthly_excess_returns = (
            sum(valid_excess_months) / len(valid_excess_months) if valid_excess_months else None
        )

        max_drawdown = None
        try:
            index_max_dd = calculate_metrics.get("index_maximum_drawdown") or {}
            start_max_dd = calculate_metrics.get("start_maximum_drawdown") or {}
            year_excess_returns = [
                int(item["year"])
                for item in (calculate_metrics.get("excess_returns") or [])
                if isinstance(item, dict)
                and item.get("year") != "all"
                and item.get("annualized_return_diff") is not None
                and item.get("annualized_return_diff") > 0
            ]
            index_year_map = {
                item["year"]: item
                for item in index_max_dd.get("year_maximum_drawdown", [])
                if isinstance(item, dict) and item.get("year") in year_excess_returns
            }
            start_year_map = {
                item["year"]: item
                for item in start_max_dd.get("year_maximum_drawdown", [])
                if isinstance(item, dict) and item.get("year") in year_excess_returns
            }
            diffs = []
            for year, index_item in index_year_map.items():
                start_item = start_year_map.get(year) or {}
                if index_item.get("drawdown") is None or start_item.get("drawdown") is None:
                    continue
                diffs.append(start_item["drawdown"] - index_item["drawdown"])
            max_drawdown = max(diffs) if diffs else None
        except Exception:
            max_drawdown = None

        total_max_drawdown = ((calculate_metrics.get("start_maximum_drawdown") or {}).get("total_maximum_drawdown") or {})

        period_text = excess_all.get("start_end_date", "")
        rows = [
            {"category": "绝对收益", "metric": "年化收益", "index_value": _fmt_percent(excess_all.get("index_annualized_return")), "model_value": _fmt_percent(excess_all.get("start_annualized_return"))},
            {"category": "绝对收益", "metric": "盈利年份百分比", "index_value": _fmt_percent(calculate_metrics.get("index_profit_annual")), "model_value": _fmt_percent(calculate_metrics.get("start_profit_annual"))},
            {"category": "绝对收益", "metric": "月盈利百分比", "index_value": _fmt_percent(index_profit_monthly_all.get("profit_monthly_percentage")), "model_value": _fmt_percent(start_profit_monthly_all.get("profit_monthly_percentage"))},
            {"category": "绝对收益", "metric": "平均月收益率", "index_value": _fmt_percent(index_sharpe_all.get("avg_monthly_return")), "model_value": _fmt_percent(start_sharpe_all.get("avg_monthly_return"))},
            {"category": "绝对收益", "metric": "月收益率波动率", "index_value": _fmt_percent(calculate_metrics.get("index_monthly_return_volatility")), "model_value": _fmt_percent(calculate_metrics.get("start_monthly_return_volatility"))},
            {"category": "相对收益", "metric": "年化超额收益", "index_value": "", "model_value": _fmt_percent(excess_all.get("annualized_return_diff"))},
            {"category": "相对收益", "metric": "跑赢年份(百分比)", "index_value": "", "model_value": _fmt_percent(calculate_metrics.get("outperform_year"))},
            {"category": "相对收益", "metric": "月超额收益胜率", "index_value": "", "model_value": _fmt_percent(monthly_excess_percentage_all.get("excess_return"))},
            {"category": "相对收益", "metric": "平均月超额", "index_value": "", "model_value": _fmt_percent(avg_monthly_excess_returns)},
            {"category": "相对收益", "metric": "月超额波动率", "index_value": "", "model_value": _fmt_percent(calculate_metrics.get("monthly_excess_volatility"))},
            {"category": "回撤", "metric": "年最大超额回撤", "index_value": "", "model_value": _fmt_percent(-max_drawdown) if max_drawdown is not None else ""},
            {"category": "回撤", "metric": "超额回撤胜率", "index_value": "", "model_value": _fmt_percent(-(calculate_metrics.get("excess_drawdown_winning_rate"))) if calculate_metrics.get("excess_drawdown_winning_rate") is not None else ""},
            {"category": "回撤", "metric": "年最大回撤", "index_value": "", "model_value": _fmt_percent(-(total_max_drawdown.get("drawdown"))) if total_max_drawdown.get("drawdown") is not None else ""},
            {"category": "回撤", "metric": "最大修复天数", "index_value": "", "model_value": str(calculate_metrics.get("start_maximum_number_of_backtest_repair_days") or "")},
            {"category": "回撤", "metric": "超额最大修复天数", "index_value": "", "model_value": str(calculate_metrics.get("excess_maximum_number_of_backtest_repair_days") or "")},
            {"category": "比率", "metric": "夏普比率", "index_value": _fmt_number(index_sharpe_all.get("sharpe_ratio")), "model_value": _fmt_number(start_sharpe_all.get("sharpe_ratio"))},
            {"category": "比率", "metric": "卡玛比率", "index_value": _fmt_number(index_kama_all.get("kama_ratio")), "model_value": _fmt_number(start_kama_all.get("kama_ratio"))},
            {"category": "比率", "metric": "所提诺比率", "index_value": _fmt_number(index_sotino_all.get("sotino_ratio")), "model_value": _fmt_number(start_sotino_all.get("sotino_ratio"))},
            {"category": "夏普", "metric": "超额夏普", "index_value": "", "model_value": _fmt_number(calculate_metrics.get("excess_sharp"))},
            {"category": "所提诺", "metric": "超额所提诺比率", "index_value": "", "model_value": _fmt_number(calculate_metrics.get("excess_of_promissory_note"))},
        ]
        return period_text, rows

    try:
        summary_df = xpl_analyzer.format_export_file_data({
            "analyze_result": calculate_metrics,
            "filename_title": model_name,
        })
    except Exception:
        return _build_fallback_rows()

    period_text = str(summary_df.iat[1, 1] or "").strip()
    rows = []
    for row_index in range(3, 23):
        category = str(summary_df.iat[row_index, 0] or "").strip()
        metric = str(summary_df.iat[row_index, 1] or "").strip()
        if not category and not metric:
            continue
        rows.append({
            "category": category,
            "metric": _normalize_metric_label(metric),
            "index_value": _normalize_display_value(summary_df.iat[row_index, 2]),
            "model_value": _normalize_display_value(summary_df.iat[row_index, 3]),
        })
    return period_text, rows


def _build_global_preview_payload(task_id):
    task = db.session.get(Task, task_id)
    if not task:
        return None
    if normalize_task_type(task.task_type) != "backtest_training":
        return None

    task_config = task.to_dict().get("config") or {}
    task_results = (
        TaskResult.query
        .options(
            load_only(
                TaskResult.id,
                TaskResult.task_id,
                TaskResult.step_index,
                TaskResult.parameters,
                TaskResult.result,
                TaskResult.success,
                TaskResult.error_message,
                TaskResult.timestamp,
            )
        )
        .filter_by(task_id=task_id)
        .order_by(TaskResult.step_index.asc(), TaskResult.timestamp.asc(), TaskResult.id.asc())
        .all()
    )

    groups = OrderedDict()
    success_count = 0
    failed_count = 0

    for task_result in task_results:
        parameters = json.loads(task_result.parameters) if task_result.parameters else {}
        year_key = str(parameters.get("year") or "未分组")
        model_name = _detect_model_name(task.name, parameters)
        group = groups.setdefault(year_key, {
            "group_key": year_key,
            "group_label": f"{year_key} 年",
            "year": year_key,
            "period": "",
            "columns": [],
            "rows": OrderedDict(),
            "failed_results": 0,
        })

        column_key = f"result_{task_result.id}"
        result_core = _extract_result_core(task_result) if task_result.success else {}
        group["columns"].append({
            "column_key": column_key,
            "result_id": task_result.id,
            "step_index": task_result.step_index,
            "header": _build_parameter_header(parameters),
            "model_name": model_name,
            "success": bool(task_result.success),
            "timestamp": task_result.timestamp.isoformat() if task_result.timestamp else None,
            "parameter_values": parameters.get("parameter") if isinstance(parameters.get("parameter"), list) else [],
            "raw_metrics": _extract_raw_sheet_metrics(result_core),
        })

        if not task_result.success:
            failed_count += 1
            group["failed_results"] += 1
            continue

        calculate_metrics = result_core.get("calculate_metrics") if isinstance(result_core, dict) else {}
        period_text, summary_rows = _extract_summary_rows(calculate_metrics, model_name)
        if not summary_rows:
            failed_count += 1
            group["failed_results"] += 1
            continue

        success_count += 1
        if period_text and not group["period"]:
            group["period"] = period_text

        for summary_row in summary_rows:
            row_key = f"{summary_row['category']}::{summary_row['metric']}"
            row = group["rows"].setdefault(row_key, {
                "category": summary_row["category"],
                "metric": summary_row["metric"],
                "index_value": summary_row["index_value"],
                "values": {},
            })
            if not row["index_value"] and summary_row["index_value"]:
                row["index_value"] = summary_row["index_value"]
            row["values"][column_key] = summary_row["model_value"]

    serialized_groups = []
    for year_key in sorted(groups.keys(), reverse=True):
        group = groups[year_key]
        ordered_rows = []
        for row in group["rows"].values():
            ordered_rows.append({
                "category": row["category"],
                "metric": row["metric"],
                "index_value": row["index_value"],
                "values": {
                    column["column_key"]: row["values"].get(column["column_key"], "")
                    for column in group["columns"]
                },
            })

        serialized_groups.append({
            "group_key": group["group_key"],
            "group_label": group["group_label"],
            "year": group["year"],
            "period": group["period"],
            "columns": group["columns"],
            "rows": ordered_rows,
            "failed_results": group["failed_results"],
            "column_count": len(group["columns"]),
        })

    return {
        "task": {
            "id": task.id,
            "name": task.name,
            "status": task.status,
            "stock_code": task_config.get("stock_code"),
            "market_type": task_config.get("market_type"),
        },
        "summary": {
            "total_results": len(task_results),
            "success_results": success_count,
            "failed_results": failed_count,
            "group_count": len(serialized_groups),
        },
        "groups": serialized_groups,
    }


def _sanitize_excel_sheet_name(name, fallback):
    raw_name = str(name or fallback or "Sheet")
    invalid_chars = set('\\/:*?[]')
    cleaned = ''.join('_' if char in invalid_chars else char for char in raw_name).strip()
    cleaned = cleaned[:31].strip() or fallback
    return cleaned


def _append_global_summary_sheet(workbook, payload, styles):
    groups = payload.get("groups") or []
    sheet = workbook.create_sheet("汇总", 0)
    if not groups:
        sheet.append(["暂无可导出的分组数据"])
        return sheet

    max_columns = 2
    header_columns = []
    for group in groups:
        columns = group.get("columns") or []
        if len(columns) > len(header_columns):
            header_columns = columns
        max_columns = max(max_columns, 2 + len(columns))

    header = ["周期", "名称"]
    for column in header_columns:
        header.append(column.get("header") or f"结果 {column.get('result_id')}")
    while len(header) < max_columns:
        header.append("")
    sheet.append(header)

    for group in groups:
        start_row = sheet.max_row + 1
        columns = group.get("columns") or []
        for metric_key, label in SUMMARY_ROW_LABELS:
            values = [group.get("year") or group.get("group_label") or "", label]
            for column in columns:
                values.append(_get_summary_derived_value(column, metric_key))
            while len(values) < max_columns:
                values.append("")
            sheet.append(values)

        end_row = sheet.max_row
        if end_row > start_row:
            sheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=end_row,
                end_column=1,
            )
    sheet.freeze_panes = "C2"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 28
    for column_index in range(3, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18

    sheet.row_dimensions[1].height = 24
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 22

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = styles["center_alignment"]
            cell.border = styles["thin_border"]
            cell.font = styles["body_font"]
            if cell.row == 1:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
            elif cell.column == 1:
                cell.font = styles["header_font"]
            elif cell.column == 2:
                cell.fill = styles["first_col_fill"]

    return sheet


def _build_global_preview_workbook(payload):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="F7E1A1")
    sub_header_fill = PatternFill("solid", fgColor="FCECC5")
    first_col_fill = PatternFill("solid", fgColor="F7E1A1")
    title_font = Font(name="Microsoft YaHei", size=12, bold=True)
    header_font = Font(name="Microsoft YaHei", size=11, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10, bold=False)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    styles = {
        "header_fill": header_fill,
        "sub_header_fill": sub_header_fill,
        "first_col_fill": first_col_fill,
        "title_font": title_font,
        "header_font": header_font,
        "body_font": body_font,
        "center_alignment": center_alignment,
        "thin_border": thin_border,
    }

    groups = payload.get("groups") or []
    if not groups:
        sheet = workbook.create_sheet("全局预览")
        sheet.append(["暂无可导出的分组数据"])
        return workbook

    _append_global_summary_sheet(workbook, payload, styles)

    used_sheet_names = set()
    for index, group in enumerate(groups, start=1):
        base_name = _sanitize_excel_sheet_name(group.get("group_label"), f"分组{index}")
        sheet_name = base_name
        suffix = 1
        while sheet_name in used_sheet_names:
            suffix += 1
            suffix_text = f"_{suffix}"
            sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
        used_sheet_names.add(sheet_name)

        sheet = workbook.create_sheet(sheet_name)
        task = payload.get("task") or {}
        stock_code = task.get("stock_code") or group.get("year") or task.get("name") or ""
        columns = group.get("columns") or []
        row_1 = [stock_code, "", ""]
        for column in columns:
            row_1.append(column.get("header") or f"结果 {column.get('result_id')}")
        sheet.append(row_1)

        header = ["指标类型", "指标", "指数"]
        for column in columns:
            model_label = column.get("model_name") or "模型"
            if not column.get("success", True):
                model_label = f"{model_label}(失败)"
            header.append(model_label)
        sheet.append(header)

        for row in group.get("rows") or []:
            values = [
                row.get("category") or "",
                row.get("metric") or "",
                row.get("index_value") or "",
            ]
            for column in columns:
                values.append((row.get("values") or {}).get(column.get("column_key"), ""))
            sheet.append(values)

        if columns:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        sheet.freeze_panes = "A3"
        width_map = {
            "A": 14,
            "B": 22,
            "C": 14,
        }
        for column_index in range(4, len(header) + 1):
            width_map[get_column_letter(column_index)] = 18
        for key, width in width_map.items():
            sheet.column_dimensions[key].width = width

        sheet.row_dimensions[1].height = 26
        sheet.row_dimensions[2].height = 24

        max_row = sheet.max_row
        max_col = sheet.max_column
        for row_index in range(1, max_row + 1):
            for col_index in range(1, max_col + 1):
                cell = sheet.cell(row=row_index, column=col_index)
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.font = body_font
                if row_index == 1:
                    cell.fill = header_fill
                    cell.font = title_font
                elif row_index == 2:
                    cell.fill = sub_header_fill
                    cell.font = header_font
                if col_index == 1 and row_index >= 2:
                    cell.fill = first_col_fill
                    if row_index == 2:
                        cell.font = header_font

        if group.get("period"):
            period_col = max_col + 1
            sheet.cell(row=1, column=period_col, value="区间")
            sheet.cell(row=2, column=period_col, value=group.get("period"))
            period_letter = get_column_letter(period_col)
            sheet.column_dimensions[period_letter].width = 24
            for row_index in (1, 2):
                cell = sheet.cell(row=row_index, column=period_col)
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.font = title_font if row_index == 1 else body_font
                cell.fill = header_fill if row_index == 1 else sub_header_fill

    return workbook


@bp.route("/api/global-preview/<task_id>", methods=["GET"])
@login_required
@permission_required('backtest:view')
def get_global_preview(task_id):
    _, error_response = _load_backtest_task_or_response(task_id, action="view")
    if error_response:
        return error_response

    payload = _build_global_preview_payload(task_id)
    if payload is None:
        return jsonify({
            "status": "error",
            "message": "任务不存在",
        }), 404

    return jsonify({
        "status": "success",
        **_sanitize_json_value(payload),
    })


@bp.route("/api/global-preview/<task_id>/export", methods=["GET"])
@login_required
@permission_required('backtest:view')
def export_global_preview(task_id):
    _, error_response = _load_backtest_task_or_response(task_id, action="view")
    if error_response:
        return error_response

    payload = _build_global_preview_payload(task_id)
    if payload is None:
        return jsonify({
            "status": "error",
            "message": "任务不存在",
        }), 404

    workbook = _build_global_preview_workbook(payload)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    task_name = (payload.get("task") or {}).get("name") or task_id
    safe_name = "".join(char if char not in '\\/:*?\"<>|' else "_" for char in str(task_name)).strip() or task_id
    filename = f"{safe_name}_global_preview.xlsx"

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
