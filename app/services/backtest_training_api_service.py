
import json
import math
import re
from collections import OrderedDict
from copy import deepcopy
from decimal import Decimal, InvalidOperation

from flask import Blueprint, current_app, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import load_only

from app.extensions import db
from app.models import Task, TaskResult, TaskResultReturn
from app.services.performance_analysis.historical_metrics import resolve_preview_metrics
from app.services.xpl_service import xpl_analyzer
from app.utils.return_series import parse_return_series_fields
from app.utils.c7_result_normalizer import (
    C7_RAW_PERCENT_CELLS,
    normalize_c7_result_metrics,
)
from app.utils.task_types import normalize_task_type




C3_PARAMETER_FIELDS = [
    ("commission", "Commission"),
    ("xm", "X Multiplier"),
    ("dbbh1", "单边保护1"),
    ("dbbh2", "单边保护2"),
    ("zlxc", "中立限仓"),
    ("zsgz", "指数跟踪"),
    ("ywf1", "一窝蜂 smoothing"),
    ("ywf2", "一窝蜂 bordering"),
]


SCIENTIFIC_NOTATION_RE = re.compile(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)[eE][+-]?\d+$")
SUMMARY_METRIC_CELL_MAP = {
    "C3": {
        "index_return": "I18",
        "return": "I15",
        "index_max_drawdown": "I20",
        "max_drawdown": "I17",
    },
    "C5": {
        "index_return": "D5",
        "return": "D2",
        "index_max_drawdown": "D7",
        "max_drawdown": "D4",
    },
    "C7": {
        "index_return": "D11",
        "return": "D8",
        "index_max_drawdown": "D13",
        "max_drawdown": "D10",
    },
}
SUMMARY_METRIC_CELL_MAP["C4"] = SUMMARY_METRIC_CELL_MAP["C5"]

SUMMARY_ROW_LABELS = [
    ("index_return", "指数回报"),
    ("return", "模型回报"),
    ("excess_return", "超额回报"),
    ("index_max_drawdown", "指数回撤"),
    ("max_drawdown", "模型回撤"),
    ("excess_drawdown", "超额回撤"),
]

def _normalize_scientific_text(text: str) -> str:
    if not SCIENTIFIC_NOTATION_RE.fullmatch(text):
        return text

    try:
        number = Decimal(text)
    except InvalidOperation:
        return text

    if not number.is_finite():
        return text

    normalized = format(number.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    if normalized in {"-0", "+0"}:
        return "0"
    return normalized




def _load_backtest_task_or_response(task_id: str, action: str = "view", result_id: int | None = None):
    task = db.session.get(Task, task_id)
    if not task:
        return None, (jsonify({
            "status": "error",
            "message": "任务不存在",
        }), 404)

    if normalize_task_type(task.task_type) != "backtest_training":
        return None, (jsonify({
            "status": "error",
            "message": "当前接口仅支持回测任务",
            "task_id": task_id,
            "task_type": task.task_type,
        }), 400)

    return task, None


def _build_zip_member_name(task_name: str | None, fallback_id: str, used_names: set[str]) -> str:
    base_name = "".join(char if char not in '\\/:*?"<>|' else "_" for char in str(task_name or "").strip())
    base_name = base_name.rstrip(" .") or fallback_id
    filename = f"{base_name}_global_preview.xlsx"
    if filename not in used_names:
        used_names.add(filename)
        return filename

    stem = filename[:-5]
    index = 2
    while True:
        candidate = f"{stem}_{index}.xlsx"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        index += 1


def _validate_batch_global_preview_task_ids(raw_task_ids):
    if not isinstance(raw_task_ids, list) or not raw_task_ids:
        return None, (jsonify({"status": "error", "message": "请选择至少一个任务"}), 400)

    task_ids = [str(task_id).strip() for task_id in raw_task_ids if str(task_id).strip()]
    task_ids = list(dict.fromkeys(task_ids))
    if not task_ids:
        return None, (jsonify({"status": "error", "message": "请选择至少一个任务"}), 400)

    return task_ids, None


def _sanitize_json_value(value):
    """Convert NaN/Infinity values into JSON-safe nulls."""
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, dict):
        return {key: _sanitize_json_value(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_value(item) for item in value]
    return value


def _strip_html_tags(value):
    return re.sub(r"<[^>]+>", "", str(value or "")).strip()


def _infer_backtest_model_version(config):
    if not isinstance(config, dict):
        return "c3"

    sheet = config.get("sheet") or {}
    title = str(sheet.get("title") or config.get("title") or "").upper()
    if "C7" in title:
        return "c7"
    if "C5" in title or "C4" in title:
        return "c5"

    parameters = config.get("parameters") or []
    first_row = parameters[0] if parameters and isinstance(parameters[0], list) else []
    if len(first_row) == 2:
        return "c5"
    return "c3"


def _is_c7_0_3_backtest_config(config):
    if not isinstance(config, dict):
        return False
    sheet = config.get("sheet") or {}
    version = str(sheet.get("c7_model_version") or config.get("c7_model_version") or "").strip().lower()
    title = str(sheet.get("title") or config.get("title") or "").upper()
    if version == "c7_0_3" or "C7.0.3" in title:
        return True
    return any(
        str(item.get("c7_model_version") or "").strip().lower() == "c7_0_3"
        for item in config.get("sheets") or []
        if isinstance(item, dict)
    )


def _resolve_c7_model_version(task_config, parameters):
    """以任务或结果明确记录的版本确定 C7 结果布局。"""
    result_version = str(parameters.get("c7_model_version") or "").strip().lower()
    if result_version in {"c7_0_2", "c7_0_3"}:
        return result_version
    if _is_c7_0_3_backtest_config(task_config):
        return "c7_0_3"
    return "c7_0_2"


def _infer_backtest_export_model_name(config):
    if not isinstance(config, dict):
        return "C3"

    sheet = config.get("sheet") or {}
    sources = (
        config.get("model_name"),
        sheet.get("title"),
        config.get("title"),
        config.get("model_version"),
    )
    for source in sources:
        title = str(source or "").upper()
        for model_name in ("C7", "C5", "C4", "C3"):
            if model_name in title:
                return model_name
    return _infer_backtest_model_version(config).upper()


def _parse_percent_like_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value if math.isfinite(value) else None

    raw = str(value).strip()
    if not raw or raw == "-":
        return None

    normalized = raw.replace(",", "").replace("$", "")
    sign = 1
    while normalized.startswith("-"):
        sign *= -1
        normalized = normalized[1:]
    try:
        if normalized.endswith("%"):
            return sign * float(normalized[:-1]) / 100
        return sign * float(normalized)
    except (TypeError, ValueError):
        return raw


def _get_task_result_return_rows(task_result, return_series_by_id=None):
    """读取单条结果的收益序列，供历史预览按需补全 V1 指标。"""
    series_id = task_result.return_series_id
    if not series_id:
        return []
    series = (
        return_series_by_id.get(series_id)
        if return_series_by_id is not None
        else db.session.get(TaskResultReturn, series_id)
    )
    return parse_return_series_fields(series) if series is not None else []


def _extract_task_result_payload(task_result, return_rows=None):
    try:
        result_payload = json.loads(task_result.result) if task_result.result else {}
    except (TypeError, json.JSONDecodeError):
        result_payload = {}

    if isinstance(result_payload, dict) and result_payload:
        value = next(
            (
                item
                for item in result_payload.values()
                if isinstance(item, dict) and (
                    "metrics_payload" in item
                    or "calculate_metrics" in item
                    or "analyze_result" in item
                )
            ),
            next((item for item in result_payload.values() if isinstance(item, dict)), {}),
        )
    else:
        value = {}
    if not isinstance(value, dict):
        return {}, {}

    calculate_metrics = resolve_preview_metrics(value, return_rows=return_rows)
    sheet_result = {
        key: item
        for key, item in value.items()
        if key not in {"metrics_payload", "calculate_metrics", "analyze_result"}
    }
    return calculate_metrics, sheet_result


def _load_backtest_task_result_or_response(task_result_id: int):
    task_result = (
        TaskResult.query
        .options(load_only(TaskResult.id, TaskResult.task_id, TaskResult.result))
        .filter(TaskResult.id == task_result_id)
        .first()
    )
    if not task_result:
        return None, None, (jsonify({
            "status": "error",
            "message": "任务结果不存在",
        }), 404)

    task, error_response = _load_backtest_task_or_response(
        task_result.task_id,
        action="view",
        result_id=task_result_id,
    )
    if error_response:
        return None, None, error_response

    return task_result, task, None


def _build_backtest_result_export_filename(task: Task, result_id: int) -> tuple[str, str]:
    filename_title = "".join(
        char if char not in '\\/:*?"<>|' else "_"
        for char in str(task.name or "").strip()
    ).rstrip(" .") or f"backtest_result_{result_id}"
    return f"{filename_title}.csv", filename_title


def _build_backtest_result_export_data(task_result: TaskResult, task: Task) -> dict:
    calculate_metrics, sheet_result = _extract_task_result_payload(
        task_result,
        return_rows=_get_task_result_return_rows(task_result),
    )
    task_config = task.to_dict().get("config") or {}
    model_name = _infer_backtest_export_model_name(task_config)
    if model_name == "C7" and not _is_c7_0_3_backtest_config(task_config):
        sheet_result = normalize_c7_result_metrics(sheet_result)
    analyze_result = {
        **calculate_metrics,
        "sheet_result": sheet_result,
        "model_name": model_name,
        "stock_code": task_config.get("stock_code"),
    }
    filename, filename_title = _build_backtest_result_export_filename(task, task_result.id)
    return {
        "filename": filename,
        "filename_title": filename_title,
        "model_name": model_name,
        "analyze_result": analyze_result,
    }


def _build_backtest_result_export_rows(export_data: dict) -> list[list[str]]:
    dataframe = xpl_analyzer.format_export_file_data(export_data)
    return [
        ["" if value is None else str(value) for value in row]
        for row in dataframe.fillna("").values.tolist()
    ]


def _extract_year_drawdown_map(section):
    if not isinstance(section, dict):
        return {}
    return {
        str(item.get("year")): item
        for item in section.get("year_maximum_drawdown", [])
        if isinstance(item, dict) and item.get("year") not in (None, "", "all")
    }


def _extract_year_sharpe_map(section):
    if not isinstance(section, dict):
        return {}

    year_map = {}
    for key, value in section.items():
        if not isinstance(value, dict):
            continue
        match = re.match(r"^year_\d+_(\d{4})$", str(key))
        if match:
            year_map[match.group(1)] = value
    return year_map


def _extract_display_year(source_window):
    raw = str(source_window or "").strip()
    if not raw:
        return ""

    if re.fullmatch(r"\d{4}-\d{4}", raw):
        end_year, start_year = raw.split("-", 1)
        return start_year or end_year

    return raw


def _build_c3_summary_rows(task_id):
    task_results = (
        TaskResult.query
        .options(
            load_only(
                TaskResult.id,
                TaskResult.task_id,
                TaskResult.step_index,
                TaskResult.parameters,
                TaskResult.result,
                TaskResult.success,
                TaskResult.timestamp,
            )
        )
        .filter_by(task_id=task_id, success=True)
        .order_by(TaskResult.step_index.asc(), TaskResult.timestamp.asc(), TaskResult.id.asc())
        .all()
    )

    rows = []

    for task_result in task_results:
        try:
            parameters = json.loads(task_result.parameters) if task_result.parameters else {}
        except (TypeError, json.JSONDecodeError):
            parameters = {}

        if not isinstance(parameters, dict):
            continue

        parameter_values = parameters.get("parameter")
        if not isinstance(parameter_values, list) or not parameter_values:
            continue

        parameter_map = {}
        for index, (field_key, _field_label) in enumerate(C3_PARAMETER_FIELDS):
            parameter_map[field_key] = parameter_values[index] if index < len(parameter_values) else None

        calculate_metrics, sheet_result = _extract_task_result_payload(
            task_result,
            return_rows=_get_task_result_return_rows(task_result),
        )
        def _safe_all_entry(items, key_name="year"):
            if not isinstance(items, list):
                return {}
            for item in items:
                if isinstance(item, dict) and str(item.get(key_name)) == "all":
                    return item
            return {}

        index_sharpe_all = (
            (calculate_metrics.get("index_sharpe_ratios") or {}).get("all")
            if isinstance(calculate_metrics.get("index_sharpe_ratios"), dict)
            else {}
        ) or {}
        start_sharpe_all = (
            (calculate_metrics.get("start_sharpe_ratios") or {}).get("all")
            if isinstance(calculate_metrics.get("start_sharpe_ratios"), dict)
            else {}
        ) or {}
        excess_returns = calculate_metrics.get("excess_returns") or []
        all_excess = next(
            (
                item for item in excess_returns
                if isinstance(item, dict) and str(item.get("year")) == "all"
            ),
            {}
        )
        index_profit_monthly_all = _safe_all_entry(
            calculate_metrics.get("index_profit_monthly")
        )
        start_profit_monthly_all = _safe_all_entry(
            calculate_metrics.get("start_profit_monthly")
        )
        index_kama_all = _safe_all_entry(calculate_metrics.get("index_kama_ratio"))
        start_kama_all = _safe_all_entry(calculate_metrics.get("start_kama_ratio"))
        index_sortino_all = _safe_all_entry(
            calculate_metrics.get("index_sortino_ratio")
        )
        start_sortino_all = _safe_all_entry(
            calculate_metrics.get("start_sortino_ratio")
        )
        monthly_excess_percentage_all = _safe_all_entry(
            calculate_metrics.get("monthly_excess_return_percentage")
        )
        monthly_excess_returns = calculate_metrics.get("monthly_excess_returns") or []
        monthly_excess_values = [
            item.get("monthly_excess_return_diff")
            for item in monthly_excess_returns
            if isinstance(item, dict)
            and item.get("monthly_excess_return_diff") is not None
        ]
        avg_monthly_excess_return = (
            sum(monthly_excess_values) / len(monthly_excess_values)
            if monthly_excess_values
            else None
        )

        source_window = str(parameters.get("year") or parameters.get("Kline_key") or "")
        year_label = _extract_display_year(source_window)
        parameter_signature = json.dumps(parameter_values, ensure_ascii=False)

        strategy_return = _parse_percent_like_value(sheet_result.get("I15"))
        index_return = _parse_percent_like_value(sheet_result.get("I18"))
        beats_index = None
        if isinstance(strategy_return, (int, float)) and isinstance(index_return, (int, float)):
            beats_index = strategy_return - index_return
        else:
            beats_index = _parse_percent_like_value(all_excess.get("annualized_return_diff"))

        strategy_max_drawdown = _parse_percent_like_value(sheet_result.get("I17"))
        index_max_drawdown = _parse_percent_like_value(sheet_result.get("I20"))
        drawdown_beats = None
        if isinstance(strategy_max_drawdown, (int, float)) and isinstance(index_max_drawdown, (int, float)):
            drawdown_beats = strategy_max_drawdown - index_max_drawdown

        rows.append({
            **parameter_map,
            "year": year_label,
            "strategy_return": strategy_return,
            "strategy_annualized": _parse_percent_like_value(sheet_result.get("I16")),
            "index_return": index_return,
            "index_annualized": _parse_percent_like_value(sheet_result.get("I19")),
            "beats_index": beats_index,
            "strategy_max_drawdown": strategy_max_drawdown,
            "index_max_drawdown": index_max_drawdown,
            "drawdown_beats": drawdown_beats,
            "fee_total": _parse_percent_like_value(sheet_result.get("I21")),
            "fee_annualized": _parse_percent_like_value(sheet_result.get("I22")),
            "year_rate": _parse_percent_like_value(sheet_result.get("I23")),
            "index_monthly_sharpe": _parse_percent_like_value(index_sharpe_all.get("sharpe_ratio")),
            "strategy_monthly_sharpe": _parse_percent_like_value(start_sharpe_all.get("sharpe_ratio")),
            "index_avg_monthly_return": _parse_percent_like_value(
                index_sharpe_all.get("avg_monthly_return")
            ),
            "strategy_avg_monthly_return": _parse_percent_like_value(
                start_sharpe_all.get("avg_monthly_return")
            ),
            "index_monthly_return_volatility": _parse_percent_like_value(
                calculate_metrics.get("index_monthly_return_volatility")
            ),
            "strategy_monthly_return_volatility": _parse_percent_like_value(
                calculate_metrics.get("start_monthly_return_volatility")
            ),
            "excess_annualized_return": _parse_percent_like_value(
                all_excess.get("annualized_return_diff")
            ),
            "outperform_year": _parse_percent_like_value(
                calculate_metrics.get("outperform_year")
            ),
            "monthly_excess_return_percentage": _parse_percent_like_value(
                monthly_excess_percentage_all.get("excess_return")
            ),
            "avg_monthly_excess_return": _parse_percent_like_value(
                avg_monthly_excess_return
            ),
            "monthly_excess_volatility": _parse_percent_like_value(
                calculate_metrics.get("monthly_excess_volatility")
            ),
            "index_profit_annual": _parse_percent_like_value(
                calculate_metrics.get("index_profit_annual")
            ),
            "strategy_profit_annual": _parse_percent_like_value(
                calculate_metrics.get("start_profit_annual")
            ),
            "index_profit_monthly_percentage": _parse_percent_like_value(
                index_profit_monthly_all.get("profit_monthly_percentage")
            ),
            "strategy_profit_monthly_percentage": _parse_percent_like_value(
                start_profit_monthly_all.get("profit_monthly_percentage")
            ),
            "index_kama_ratio": _parse_percent_like_value(
                index_kama_all.get("kama_ratio")
            ),
            "strategy_kama_ratio": _parse_percent_like_value(
                start_kama_all.get("kama_ratio")
            ),
            "index_sortino_ratio": _parse_percent_like_value(
                index_sortino_all.get("sortino_ratio")
            ),
            "strategy_sortino_ratio": _parse_percent_like_value(
                start_sortino_all.get("sortino_ratio")
            ),
            "excess_sharpe": _parse_percent_like_value(
                calculate_metrics.get("excess_sharpe")
            ),
            "excess_sortino": _parse_percent_like_value(
                calculate_metrics.get("excess_sortino")
            ),
            "excess_drawdown_winning_rate": _parse_percent_like_value(
                calculate_metrics.get("excess_drawdown_winning_rate")
            ),
            "index_maximum_number_of_backtest_repair_days": _parse_percent_like_value(
                calculate_metrics.get("index_maximum_number_of_backtest_repair_days")
            ),
            "strategy_maximum_number_of_backtest_repair_days": _parse_percent_like_value(
                calculate_metrics.get("start_maximum_number_of_backtest_repair_days")
            ),
            "excess_maximum_number_of_backtest_repair_days": _parse_percent_like_value(
                calculate_metrics.get("excess_maximum_number_of_backtest_repair_days")
            ),
            "date_range": all_excess.get("start_end_date"),
            "source_window": source_window,
            "task_result_id": task_result.id,
            "step_index": task_result.step_index,
            "timestamp": task_result.timestamp.isoformat() if task_result.timestamp else None,
            "parameter_signature": parameter_signature,
        })

    rows.sort(
        key=lambda item: (
            item.get("parameter_signature") or "",
            -(int(item.get("year")) if str(item.get("year", "")).isdigit() else -9999),
            item.get("step_index") or 0,
        )
    )

    parameter_group_count = len({row["parameter_signature"] for row in rows})
    for row in rows:
        row.pop("parameter_signature", None)

    return rows, parameter_group_count



def _build_parameter_header(parameters):
    parameter_values = parameters.get("parameter")
    if not isinstance(parameter_values, list) or not parameter_values:
        c7_values = [
            parameters.get("A1", parameters.get("xm")),
            parameters.get("B1", parameters.get("ml")),
        ]
        c7_values = [value for value in c7_values if value not in (None, "")]
        return " / ".join(str(value) for value in c7_values) or "未命名参数"

    if len(parameter_values) == 2:
        labels = ["xm", "ml"]
    else:
        labels = [f"参数{i + 1}" for i in range(len(parameter_values))]

    parts = []
    for _, value in zip(labels, parameter_values):
        parts.append(f"{value}")
    return " / ".join(parts)


def _extract_result_core(task_result):
    payload = task_result.to_dict().get("result") or {}
    if not isinstance(payload, dict) or not payload:
        return {}
    first_value = next(iter(payload.values()), {})
    return first_value if isinstance(first_value, dict) else {}


def _detect_model_name(task_name, parameters):
    upper_name = (task_name or "").upper()
    for model_name in ("C7", "C5", "C4", "C3"):
        if model_name in upper_name:
            return model_name

    parameter_values = parameters.get("parameter")
    if isinstance(parameter_values, list) and len(parameter_values) == 2:
        return "C5"
    return "C3"


def _detect_global_preview_model_name(task, parameters):
    """C 系列页面任务类型已明确时，不依赖可变的任务名称推断模型。"""
    fixed_model_names = {
        "google_sheet": "C3",
        "google_sheet_c4": "C4",
        "google_sheet_c5": "C5",
        "google_sheet_c7": "C7",
    }
    task_type = normalize_task_type(task.task_type)
    return fixed_model_names.get(task_type) or _detect_model_name(task.name, parameters)


def _extract_raw_sheet_metrics(result_core):
    if not isinstance(result_core, dict):
        return {}
    return {
        str(key): value
        for key, value in result_core.items()
        if key not in {"metrics_payload", "calculate_metrics", "analyze_result"}
    }


def _normalize_summary_numeric_value(value):
    parsed = _parse_percent_like_value(value)
    if isinstance(parsed, (int, float)):
        return parsed if math.isfinite(parsed) else None
    return None


def _format_summary_value(value):
    parsed = _parse_percent_like_value(value)
    if isinstance(parsed, (int, float)):
        if not math.isfinite(parsed):
            return ""
        return f"{parsed:.2%}"
    if parsed is None:
        return ""
    return _normalize_scientific_text(str(parsed).strip())


def _get_summary_raw_metric(column, metric_key):
    model_name = str(column.get("model_name") or "C3").upper()
    if model_name == "C7" and column.get("c7_model_version") == "c7_0_3":
        cell_map = SUMMARY_METRIC_CELL_MAP["C5"]
    else:
        cell_map = SUMMARY_METRIC_CELL_MAP.get(model_name, SUMMARY_METRIC_CELL_MAP["C3"])
    cell_key = cell_map.get(metric_key)
    raw_metrics = column.get("raw_metrics") or {}
    value = raw_metrics.get(cell_key) if cell_key else None
    if (
        model_name == "C7"
        and column.get("c7_model_version") != "c7_0_3"
        and cell_key in C7_RAW_PERCENT_CELLS
        and value not in (None, "")
        and not str(value).strip().endswith("%")
    ):
        return normalize_c7_result_metrics({cell_key: value}).get(cell_key)
    return value


def _get_excess_return_from_calculate_metrics(calculate_metrics):
    if not isinstance(calculate_metrics, dict):
        return None

    direct_value = _normalize_summary_numeric_value(
        calculate_metrics.get("annualized_return_diff")
    )
    if direct_value is not None:
        return direct_value

    excess_returns = calculate_metrics.get("excess_returns") or []
    if not isinstance(excess_returns, list):
        return None
    all_period = next(
        (
            item for item in excess_returns
            if isinstance(item, dict) and str(item.get("year")).lower() == "all"
        ),
        None,
    )
    return _normalize_summary_numeric_value(
        (all_period or {}).get("annualized_return_diff")
    )


def _get_summary_derived_value(column, metric_key, calculate_metrics=None):
    if metric_key == "excess_return":
        left = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "return")
        )
        right = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "index_return")
        )
        if left is not None and right is not None:
            return f"{left - right:.2%}"

        calculated_value = _get_excess_return_from_calculate_metrics(calculate_metrics)
        return f"{calculated_value:.2%}" if calculated_value is not None else ""

    if metric_key == "excess_drawdown":
        left = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "max_drawdown")
        )
        right = _normalize_summary_numeric_value(
            _get_summary_raw_metric(column, "index_max_drawdown")
        )
        if left is None or right is None:
            return ""
        return f"{left - right:.2%}"

    return _format_summary_value(_get_summary_raw_metric(column, metric_key))


def _negative_percent_display(value):
    parsed = _parse_percent_like_value(value)
    if not isinstance(parsed, (int, float)) or not math.isfinite(parsed):
        return ""
    if parsed == 0:
        return "0.00%"
    return f"{-abs(parsed):.2%}"


def _percent_display(value):
    parsed = _parse_percent_like_value(value)
    if not isinstance(parsed, (int, float)) or not math.isfinite(parsed):
        return ""
    return f"{0 if parsed == 0 else parsed:.2%}"


def _max_yearly_repair_days(yearly_repair_days):
    if not isinstance(yearly_repair_days, dict):
        return None
    values = [
        value for value in yearly_repair_days.values()
        if isinstance(value, (int, float)) and not isinstance(value, bool)
        and math.isfinite(value)
    ]
    return max(values) if values else None


def _metric_year_key(value):
    text = str(value if value is not None else "").strip()
    if not text or text.lower() == "all":
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    return str(int(number)) if number.is_integer() else text


def _derive_year_max_excess_drawdown(calculate_metrics):
    excess_returns = [
        (_metric_year_key(item.get("year")), _parse_percent_like_value(item.get("annualized_return_diff")))
        for item in calculate_metrics.get("excess_returns") or []
        if isinstance(item, dict)
    ]
    annual_excess_returns = [
        (year, diff)
        for year, diff in excess_returns
        if year
    ]
    if not annual_excess_returns:
        return None

    excess_years = {
        year
        for year, annualized_return_diff in annual_excess_returns
        if isinstance(annualized_return_diff, (int, float))
        and math.isfinite(annualized_return_diff)
        and annualized_return_diff > 0
    }
    if not excess_years:
        return 0.0

    index_max_dd = calculate_metrics.get("index_maximum_drawdown") or {}
    start_max_dd = calculate_metrics.get("start_maximum_drawdown") or {}
    index_year_map = {
        year: item
        for item in index_max_dd.get("year_maximum_drawdown", [])
        if isinstance(item, dict)
        for year in [_metric_year_key(item.get("year"))]
        if year in excess_years
    }
    start_year_map = {
        year: item
        for item in start_max_dd.get("year_maximum_drawdown", [])
        if isinstance(item, dict)
        for year in [_metric_year_key(item.get("year"))]
        if year in excess_years
    }

    diffs = []
    for year, index_item in index_year_map.items():
        start_item = start_year_map.get(year) or {}
        index_drawdown = _parse_percent_like_value(index_item.get("drawdown"))
        start_drawdown = _parse_percent_like_value(start_item.get("drawdown"))
        if not isinstance(index_drawdown, (int, float)) or not isinstance(start_drawdown, (int, float)):
            continue
        if not math.isfinite(index_drawdown) or not math.isfinite(start_drawdown):
            continue
        diffs.append(start_drawdown - index_drawdown)
    return max(diffs) if diffs else None


def _format_excel_data_cell(cell):
    value = cell.value
    if not isinstance(value, str):
        return

    text = value.strip()
    if not text.endswith("%"):
        return

    parsed = _parse_percent_like_value(text)
    if not isinstance(parsed, (int, float)) or not math.isfinite(parsed):
        return

    cell.value = 0 if parsed == 0 else parsed
    cell.number_format = "0.00%"


def _with_excess_return_preview_row(summary_rows, column, calculate_metrics=None):
    if not summary_rows:
        return summary_rows
    if any(
        row.get("category") == "绝对收益" and row.get("metric") == "超额回报"
        for row in summary_rows
    ):
        return summary_rows

    excess_return_row = {
        "category": "绝对收益",
        "metric": "超额回报",
        "index_value": "",
        "model_value": _get_summary_derived_value(
            column, "excess_return", calculate_metrics
        ),
    }
    rows = []
    inserted = False
    for row in summary_rows:
        if (
            not inserted
            and row.get("category") == "绝对收益"
            and row.get("metric") == "年化收益"
        ):
            rows.append(excess_return_row)
            inserted = True
        rows.append(row)

    if not inserted:
        rows.insert(0, excess_return_row)
    return rows


def _extract_summary_rows(calculate_metrics, model_name):
    if not isinstance(calculate_metrics, dict) or not calculate_metrics:
        return "", []
    calculate_metrics = _normalize_calculate_metrics_years_for_xpl_export(calculate_metrics)

    def _normalize_metric_label(label):
        text = str(label or "").strip()
        text = text.replace("（", "(").replace("）", ")")
        metric_aliases = {
            "跑赢年份(百分比)": "跑赢年份(百分比)",
            "跑赢年份(百分比 )": "跑赢年份(百分比)",
            "超额最大修复天数": "超额最大修复天数",
            "最大修复天数": "最大修复天数",
            "索提诺比率": "索提诺比率",
            "超额索提诺比率": "超额索提诺比率",
        }
        return metric_aliases.get(text, text)

    def _normalize_display_value(value):
        text = str(value or "").strip()
        if not text:
            return ""
        while text.startswith("--"):
            text = text[1:]
        return _normalize_scientific_text(text)

    def _normalize_negative_display_value(metric, value):
        if metric == "年最大回撤":
            return _negative_percent_display(value)
        return _normalize_display_value(value)

    def _fmt_percent(value):
        if value is None or not math.isfinite(value):
            return ""
        return f"{value:.2%}"

    def _fmt_number(value):
        if value is None or not math.isfinite(value):
            return ""
        return f"{value:.2f}".rstrip("0").rstrip(".")

    def _safe_all_entry(items, key_name):
        if not isinstance(items, list):
            return {}
        for item in items:
            if isinstance(item, dict) and str(item.get(key_name)) == "all":
                return item
        return {}

    def _build_fallback_rows():
        excess_all = _safe_all_entry(calculate_metrics.get("excess_returns"), "year")
        index_profit_monthly_all = _safe_all_entry(calculate_metrics.get("index_profit_monthly"), "year")
        start_profit_monthly_all = _safe_all_entry(calculate_metrics.get("start_profit_monthly"), "year")
        index_kama_all = _safe_all_entry(calculate_metrics.get("index_kama_ratio"), "year")
        start_kama_all = _safe_all_entry(calculate_metrics.get("start_kama_ratio"), "year")
        index_sortino_all = _safe_all_entry(calculate_metrics.get("index_sortino_ratio"), "year")
        start_sortino_all = _safe_all_entry(calculate_metrics.get("start_sortino_ratio"), "year")
        monthly_excess_percentage_all = _safe_all_entry(
            calculate_metrics.get("monthly_excess_return_percentage"), "year"
        )
        index_sharpe_all = (calculate_metrics.get("index_sharpe_ratios") or {}).get("all") or {}
        start_sharpe_all = (calculate_metrics.get("start_sharpe_ratios") or {}).get("all") or {}

        monthly_excess_returns = calculate_metrics.get("monthly_excess_returns") or []
        valid_excess_months = [
            item.get("monthly_excess_return_diff")
        for item in monthly_excess_returns
            if isinstance(item, dict) and item.get("monthly_excess_return_diff") is not None
        ]
        avg_monthly_excess_returns = (
            sum(valid_excess_months) / len(valid_excess_months) if valid_excess_months else None
        )

        max_drawdown = _derive_year_max_excess_drawdown(calculate_metrics)

        total_max_drawdown = ((calculate_metrics.get("start_maximum_drawdown") or {}).get("total_maximum_drawdown") or {})
        year_index_max_repair_days = _max_yearly_repair_days(
            calculate_metrics.get("year_index_yearly_max_repair_days")
        )
        year_start_max_repair_days = _max_yearly_repair_days(
            calculate_metrics.get("year_start_yearly_max_repair_days")
        )

        period_text = excess_all.get("start_end_date", "")
        rows = [
            {"category": "绝对收益", "metric": "年化收益", "index_value": _fmt_percent(excess_all.get("index_annualized_return")), "model_value": _fmt_percent(excess_all.get("start_annualized_return"))},
            {"category": "绝对收益", "metric": "盈利年份百分比", "index_value": _fmt_percent(calculate_metrics.get("index_profit_annual")), "model_value": _fmt_percent(calculate_metrics.get("start_profit_annual"))},
            {"category": "绝对收益", "metric": "月盈利百分比", "index_value": _fmt_percent(index_profit_monthly_all.get("profit_monthly_percentage")), "model_value": _fmt_percent(start_profit_monthly_all.get("profit_monthly_percentage"))},
            {"category": "绝对收益", "metric": "平均月收益率", "index_value": _fmt_percent(index_sharpe_all.get("avg_monthly_return")), "model_value": _fmt_percent(start_sharpe_all.get("avg_monthly_return"))},
            {"category": "绝对收益", "metric": "月收益率波动率", "index_value": _fmt_percent(calculate_metrics.get("index_monthly_return_volatility")), "model_value": _fmt_percent(calculate_metrics.get("start_monthly_return_volatility"))},
            {"category": "相对收益", "metric": "年化超额收益", "index_value": "", "model_value": _fmt_percent(excess_all.get("annualized_return_diff"))},
            {"category": "相对收益", "metric": "跑赢年份(百分比)", "index_value": "", "model_value": _fmt_percent(calculate_metrics.get("outperform_year"))},
            {"category": "相对收益", "metric": "月超额收益胜率", "index_value": "", "model_value": _fmt_percent(monthly_excess_percentage_all.get("excess_return"))},
            {"category": "相对收益", "metric": "平均月超额", "index_value": "", "model_value": _fmt_percent(avg_monthly_excess_returns)},
            {"category": "相对收益", "metric": "月超额波动率", "index_value": "", "model_value": _fmt_percent(calculate_metrics.get("monthly_excess_volatility"))},
            {"category": "回撤", "metric": "年最大超额回撤", "index_value": "", "model_value": _percent_display(max_drawdown) if max_drawdown is not None else ""},
            {"category": "回撤", "metric": "超额回撤胜率", "index_value": "", "model_value": _percent_display(calculate_metrics.get("excess_drawdown_winning_rate")) if calculate_metrics.get("excess_drawdown_winning_rate") is not None else ""},
            {"category": "回撤", "metric": "年最大回撤", "index_value": "", "model_value": _negative_percent_display(total_max_drawdown.get("drawdown")) if total_max_drawdown.get("drawdown") is not None else ""},
            {"category": "回撤", "metric": "最大修复天数", "index_value": "", "model_value": str(calculate_metrics.get("start_maximum_number_of_backtest_repair_days") or "")},
            {"category": "回撤", "metric": "超额最大修复天数", "index_value": "", "model_value": str(calculate_metrics.get("excess_maximum_number_of_backtest_repair_days") or "")},
            {"category": "回撤", "metric": "年最大回测修复天数", "index_value": str(year_index_max_repair_days) if year_index_max_repair_days is not None else "", "model_value": str(year_start_max_repair_days) if year_start_max_repair_days is not None else ""},
            {"category": "比率", "metric": "夏普比率", "index_value": _fmt_number(index_sharpe_all.get("sharpe_ratio")), "model_value": _fmt_number(start_sharpe_all.get("sharpe_ratio"))},
            {"category": "比率", "metric": "卡玛比率", "index_value": _fmt_number(index_kama_all.get("kama_ratio")), "model_value": _fmt_number(start_kama_all.get("kama_ratio"))},
            {"category": "比率", "metric": "索提诺比率", "index_value": _fmt_number(index_sortino_all.get("sortino_ratio")), "model_value": _fmt_number(start_sortino_all.get("sortino_ratio"))},
            {"category": "夏普", "metric": "超额夏普", "index_value": "", "model_value": _fmt_number(calculate_metrics.get("excess_sharpe"))},
            {"category": "索提诺", "metric": "超额索提诺比率", "index_value": "", "model_value": _fmt_number(calculate_metrics.get("excess_sortino"))},
        ]
        return period_text, rows

    # 全局预览仅需要固定的 20 项汇总指标。直接读取已有计算结果，避免每个
    # TaskResult 都创建一次 Pandas DataFrame（多股票导出时这是主要 CPU 开销）。
    return _build_fallback_rows()


def _normalize_calculate_metrics_years_for_xpl_export(calculate_metrics):
    normalized = deepcopy(calculate_metrics)

    def normalize_year(value):
        text = str(value if value is not None else "").strip()
        if not text or text.lower() == "all":
            return value
        try:
            number = float(text)
        except ValueError:
            return value
        return int(number) if number.is_integer() else value

    for item in normalized.get("excess_returns") or []:
        if isinstance(item, dict):
            item["year"] = normalize_year(item.get("year"))
    for drawdown_key in ("index_maximum_drawdown", "start_maximum_drawdown"):
        drawdown = normalized.get(drawdown_key)
        if not isinstance(drawdown, dict):
            continue
        for item in drawdown.get("year_maximum_drawdown") or []:
            if isinstance(item, dict):
                item["year"] = normalize_year(item.get("year"))
    return normalized


def _query_global_preview_results(task_id, result_ids=None):
    """按主键精确读取结果，避免切换分组时扫描整个任务的大 JSON。"""
    query = (
        TaskResult.query
        .options(
            load_only(
                TaskResult.id,
                TaskResult.task_id,
                TaskResult.step_index,
                TaskResult.parameters,
                TaskResult.result,
                TaskResult.return_series_id,
                TaskResult.success,
                TaskResult.error_message,
                TaskResult.timestamp,
            )
        )
        .filter_by(task_id=task_id)
        .order_by(TaskResult.step_index.asc(), TaskResult.timestamp.asc(), TaskResult.id.asc())
    )
    if result_ids is not None:
        query = query.filter(TaskResult.id.in_(result_ids))
    return query.all()


def _build_global_preview_payload_from_results(task, task_results):
    """将单产品 TaskResult 转换为全局预览使用的表格数据格式。

    单产品页面以“股票 + 年份区间”分组、以每条成功结果作为动态列；
    行指标来自 XPL 摘要格式化逻辑，缺失结果仍保留列以便用户识别失败步骤。
    """
    task_config = task.to_dict().get("config") or {}
    return_series_by_id = {
        series.id: series
        for series in TaskResultReturn.query.filter(
            TaskResultReturn.id.in_({
                item.return_series_id for item in task_results if item.return_series_id
            })
        ).all()
    }

    groups = OrderedDict()
    success_count = 0
    failed_count = 0

    # 先建立分组和动态列，再填充指标行，保证失败结果也能在预览中占位。
    for task_result in task_results:
        parameters = json.loads(task_result.parameters) if task_result.parameters else {}
        year_key = str(parameters.get("year") or parameters.get("Kline_key") or "未分组")
        stock_code = str(
            parameters.get("stock_code") or task_config.get("stock_code") or "未命名股票"
        ).strip().upper() or "未命名股票"
        group_key = f"{stock_code}::{year_key}"
        group = groups.setdefault(group_key, {
            "group_key": group_key,
            "group_label": f"{year_key} 年",
            "stock_code": stock_code,
            "year": year_key,
            "period": "",
            "columns": [],
            "rows": OrderedDict(),
            "failed_results": 0,
        })

        column_key = f"result_{task_result.id}"
        result_core = _extract_result_core(task_result) if task_result.success else {}
        model_name = _detect_global_preview_model_name(task, parameters)
        column = {
            "column_key": column_key,
            "result_id": task_result.id,
            "step_index": task_result.step_index,
            "header": _build_parameter_header(parameters),
            "model_name": model_name,
            "c7_model_version": (
                _resolve_c7_model_version(task_config, parameters)
                if model_name == "C7" else None
            ),
            "success": bool(task_result.success),
            "timestamp": task_result.timestamp.isoformat() if task_result.timestamp else None,
            "parameter_values": parameters.get("parameter") if isinstance(parameters.get("parameter"), list) else [],
            "raw_metrics": _extract_raw_sheet_metrics(result_core),
        }
        group["columns"].append(column)

        if not task_result.success:
            failed_count += 1
            group["failed_results"] += 1
            continue

        # 统一存储载荷与旧别名优先；缺新预览字段时仅用当前分组已批量加载的收益序列临时补齐。
        calculate_metrics = resolve_preview_metrics(
            result_core,
            return_rows=_get_task_result_return_rows(task_result, return_series_by_id),
        )
        period_text, summary_rows = _extract_summary_rows(calculate_metrics, model_name)
        summary_rows = _with_excess_return_preview_row(
            summary_rows, column, calculate_metrics
        )
        if not summary_rows:
            failed_count += 1
            group["failed_results"] += 1
            continue

        success_count += 1
        if period_text and not group["period"]:
            group["period"] = period_text

        # 同一指标跨多个结果列合并为一行；没有该列值时使用空字符串而非误报为 0。
        for summary_row in summary_rows:
            row_key = f"{summary_row['category']}::{summary_row['metric']}"
            row = group["rows"].setdefault(row_key, {
                "category": summary_row["category"],
                "metric": summary_row["metric"],
                "index_value": summary_row["index_value"],
                "values": {},
            })
            if not row["index_value"] and summary_row["index_value"]:
                row["index_value"] = summary_row["index_value"]
            row["values"][column_key] = summary_row["model_value"]

    # 序列化阶段固定列顺序，并把内部 OrderedDict 转成前端可直接消费的 JSON 结构。
    serialized_groups = []
    for group_key in sorted(groups.keys(), reverse=True):
        group = groups[group_key]
        ordered_rows = []
        for row in group["rows"].values():
            ordered_rows.append({
                "category": row["category"],
                "metric": row["metric"],
                "index_value": row["index_value"],
                "values": {
                    column["column_key"]: row["values"].get(column["column_key"], "")
                    for column in group["columns"]
                },
            })

        serialized_groups.append({
            "group_key": group["group_key"],
            "group_label": group["group_label"],
            "stock_code": group["stock_code"],
            "year": group["year"],
            "period": group["period"],
            "columns": group["columns"],
            "rows": ordered_rows,
            "failed_results": group["failed_results"],
            "column_count": len(group["columns"]),
        })

    return {
        "task": {
            "id": task.id,
            "name": task.name,
            "status": task.status,
            "stock_code": task_config.get("stock_code"),
            "market_type": task_config.get("market_type"),
        },
        "summary": {
            "total_results": len(task_results),
            "success_results": success_count,
            "failed_results": failed_count,
            "group_count": len(serialized_groups),
            "stock_count": len({group["stock_code"] for group in serialized_groups}),
        },
        "groups": serialized_groups,
    }


def _build_global_preview_payload(task_id):
    """兼容全量调用；页面首屏不应使用该函数。"""
    task = db.session.get(Task, task_id)
    if not task or normalize_task_type(task.task_type) not in {
        "backtest_training", "google_sheet", "google_sheet_c4", "google_sheet_c5", "google_sheet_c7",
    }:
        return None
    return _build_global_preview_payload_from_results(
        task, _query_global_preview_results(task_id)
    )


def _build_global_preview_initial_payload(task_id):
    """首屏仅加载轻量参数索引，并预加载用户默认会看到的一个分组。"""
    task = db.session.get(Task, task_id)
    if not task:
        return None
    task_config = task.to_dict().get("config") or {}
    metadata_rows = (
        db.session.query(
            TaskResult.id, TaskResult.parameters, TaskResult.success, TaskResult.step_index
        )
        .filter(TaskResult.task_id == task_id)
        .order_by(TaskResult.step_index.asc(), TaskResult.id.asc())
        .all()
    )

    items = []
    is_c7_0_3 = _is_c7_0_3_backtest_config(task_config)
    for row in metadata_rows:
        parameters = json.loads(row.parameters) if row.parameters else {}
        model_name = _detect_global_preview_model_name(task, parameters)
        if model_name == "C7" and _resolve_c7_model_version(task_config, parameters) == "c7_0_3":
            is_c7_0_3 = True
        items.append({
            "id": row.id,
            "parameters": parameters,
            "success": bool(row.success),
            "step_index": row.step_index,
        })

    # 只有 C7.0.3 按股票分组；其余 C 系列维持原有按回测年份/区间查看的习惯。
    group_mode = "stock" if is_c7_0_3 else "year"
    groups = OrderedDict()
    for item in items:
        parameters = item["parameters"]
        if group_mode == "stock":
            key = str(parameters.get("stock_code") or task_config.get("stock_code") or "未命名股票").strip().upper()
        else:
            key = str(parameters.get("year") or parameters.get("Kline_key") or "未分组")
        group = groups.setdefault(key, {"key": key, "label": key, "result_ids": [], "items": []})
        group["result_ids"].append(item["id"])
        group["items"].append(item)

    serialized_groups = list(groups.values())
    default_group = serialized_groups[0] if serialized_groups else None
    default_ids = default_group["result_ids"] if default_group else []
    return {
        "group_mode": group_mode,
        "groups": serialized_groups,
        "default_group_key": default_group["key"] if default_group else "",
        "preview": _build_global_preview_payload_from_results(
            task, _query_global_preview_results(task_id, default_ids)
        ),
    }


def _build_global_preview_group_payload(task_id, result_ids):
    """结果 ID 必须同时受 task_id 约束，防止跨任务读取。"""
    task = db.session.get(Task, task_id)
    if not task:
        return None
    safe_ids = [int(item) for item in result_ids if str(item).isdigit()]
    return _build_global_preview_payload_from_results(
        task, _query_global_preview_results(task_id, safe_ids)
    )


def get_global_preview_result_ids_by_stock(task_id):
    """导出用的轻量索引：先分股票，再逐股票读取完整结果生成文件。"""
    task = db.session.get(Task, task_id)
    if not task:
        return None, []
    task_config = task.to_dict().get("config") or {}
    rows = (
        db.session.query(TaskResult.id, TaskResult.parameters)
        .filter(TaskResult.task_id == task_id)
        .order_by(TaskResult.step_index.asc(), TaskResult.id.asc())
        .all()
    )
    stock_groups = OrderedDict()
    for row in rows:
        parameters = json.loads(row.parameters) if row.parameters else {}
        stock_code = str(
            parameters.get("stock_code") or task_config.get("stock_code") or "未命名股票"
        ).strip().upper() or "未命名股票"
        stock_groups.setdefault(stock_code, []).append(row.id)
    return task, list(stock_groups.items())


def split_global_preview_payload_by_stock(payload):
    """按股票代码拆分全局预览载荷。"""
    grouped_payloads = OrderedDict()
    for group in payload.get("groups") or []:
        stock_code = str(
            group.get("stock_code")
            or (payload.get("task") or {}).get("stock_code")
            or "未命名股票"
        ).strip().upper() or "未命名股票"
        grouped_payloads.setdefault(stock_code, []).append(group)

    return [
        (
            stock_code,
            {
                **payload,
                "task": {**(payload.get("task") or {}), "stock_code": stock_code},
                "summary": {
                    **(payload.get("summary") or {}),
                    "group_count": len(groups),
                    "stock_count": 1,
                },
                "groups": groups,
            },
        )
        for stock_code, groups in grouped_payloads.items()
    ]


def _sanitize_excel_sheet_name(name, fallback):
    raw_name = str(name or fallback or "Sheet")
    invalid_chars = set('\\/:*?[]')
    cleaned = ''.join('_' if char in invalid_chars else char for char in raw_name).strip()
    cleaned = cleaned[:31].strip() or fallback
    return cleaned


def _append_global_summary_sheet(workbook, payload, styles):
    groups = payload.get("groups") or []
    sheet = workbook.create_sheet("汇总", 0)
    if not groups:
        sheet.append(["暂无可导出的分组数据"])
        return sheet

    max_columns = 2
    header_columns = []
    for group in groups:
        columns = group.get("columns") or []
        if len(columns) > len(header_columns):
            header_columns = columns
        max_columns = max(max_columns, 2 + len(columns))

    header = ["周期", "名称"]
    for column in header_columns:
        header.append(column.get("header") or f"结果 {column.get('result_id')}")
    while len(header) < max_columns:
        header.append("")
    sheet.append(header)

    for group in groups:
        start_row = sheet.max_row + 1
        columns = group.get("columns") or []
        for metric_key, label in SUMMARY_ROW_LABELS:
            values = [group.get("year") or group.get("group_label") or "", label]
            for column in columns:
                values.append(_get_summary_derived_value(column, metric_key))
            while len(values) < max_columns:
                values.append("")
            sheet.append(values)

        end_row = sheet.max_row
        if end_row > start_row:
            sheet.merge_cells(
                start_row=start_row,
                start_column=1,
                end_row=end_row,
                end_column=1,
            )
    sheet.freeze_panes = "C2"
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 28
    for column_index in range(3, max_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 18

    sheet.row_dimensions[1].height = 24
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 22

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = styles["center_alignment"]
            cell.border = styles["thin_border"]
            cell.font = styles["body_font"]
            if cell.row == 1:
                cell.fill = styles["header_fill"]
                cell.font = styles["header_font"]
            elif cell.column == 1:
                cell.font = styles["header_font"]
            elif cell.column == 2:
                cell.fill = styles["first_col_fill"]
            if cell.row >= 2 and cell.column >= 3:
                _format_excel_data_cell(cell)

    return sheet


def _build_global_preview_workbook(payload):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="F7E1A1")
    sub_header_fill = PatternFill("solid", fgColor="FCECC5")
    first_col_fill = PatternFill("solid", fgColor="F7E1A1")
    title_font = Font(name="Microsoft YaHei", size=12, bold=True)
    header_font = Font(name="Microsoft YaHei", size=11, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10, bold=False)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    styles = {
        "header_fill": header_fill,
        "sub_header_fill": sub_header_fill,
        "first_col_fill": first_col_fill,
        "title_font": title_font,
        "header_font": header_font,
        "body_font": body_font,
        "center_alignment": center_alignment,
        "thin_border": thin_border,
    }

    groups = payload.get("groups") or []
    if not groups:
        sheet = workbook.create_sheet("全局预览")
        sheet.append(["暂无可导出的分组数据"])
        return workbook

    _append_global_summary_sheet(workbook, payload, styles)

    used_sheet_names = set()
    for index, group in enumerate(groups, start=1):
        base_name = _sanitize_excel_sheet_name(group.get("group_label"), f"分组{index}")
        sheet_name = base_name
        suffix = 1
        while sheet_name in used_sheet_names:
            suffix += 1
            suffix_text = f"_{suffix}"
            sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
        used_sheet_names.add(sheet_name)

        sheet = workbook.create_sheet(sheet_name)
        task = payload.get("task") or {}
        stock_code = task.get("stock_code") or group.get("year") or task.get("name") or ""
        columns = group.get("columns") or []
        row_1 = [stock_code, "", ""]
        for column in columns:
            row_1.append(column.get("header") or f"结果 {column.get('result_id')}")
        sheet.append(row_1)

        header = ["指标类型", "指标", "指数"]
        for column in columns:
            model_label = column.get("model_name") or "模型"
            if not column.get("success", True):
                model_label = f"{model_label}(失败)"
            header.append(model_label)
        sheet.append(header)

        for row in group.get("rows") or []:
            values = [
                row.get("category") or "",
                row.get("metric") or "",
                row.get("index_value") or "",
            ]
            for column in columns:
                values.append((row.get("values") or {}).get(column.get("column_key"), ""))
            sheet.append(values)

        if columns:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        sheet.freeze_panes = "A3"
        width_map = {
            "A": 14,
            "B": 22,
            "C": 14,
        }
        for column_index in range(4, len(header) + 1):
            width_map[get_column_letter(column_index)] = 18
        for key, width in width_map.items():
            sheet.column_dimensions[key].width = width

        sheet.row_dimensions[1].height = 26
        sheet.row_dimensions[2].height = 24

        max_row = sheet.max_row
        max_col = sheet.max_column
        for row_index in range(1, max_row + 1):
            for col_index in range(1, max_col + 1):
                cell = sheet.cell(row=row_index, column=col_index)
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.font = body_font
                if row_index == 1:
                    cell.fill = header_fill
                    cell.font = title_font
                elif row_index == 2:
                    cell.fill = sub_header_fill
                    cell.font = header_font
                if col_index == 1 and row_index >= 2:
                    cell.fill = first_col_fill
                    if row_index == 2:
                        cell.font = header_font
                if row_index >= 3 and col_index >= 3:
                    _format_excel_data_cell(cell)

        if group.get("period"):
            period_col = max_col + 1
            sheet.cell(row=1, column=period_col, value="区间")
            sheet.cell(row=2, column=period_col, value=group.get("period"))
            period_letter = get_column_letter(period_col)
            sheet.column_dimensions[period_letter].width = 24
            for row_index in (1, 2):
                cell = sheet.cell(row=row_index, column=period_col)
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.font = title_font if row_index == 1 else body_font
                cell.fill = header_fill if row_index == 1 else sub_header_fill

    return workbook

