from __future__ import annotations

import ast
import operator
import re
import uuid
from pathlib import Path
from typing import Any

import openpyxl
from flask import current_app
from openpyxl.utils import column_index_from_string
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename


class BacktestExcelService:
    """Import fixed-template backtest tasks from Excel files."""

    ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}

    def storage_dir(self) -> Path:
        return Path(current_app.root_path).parent / "data" / "backtest_excel"

    def import_uploaded_excel(self, file: FileStorage) -> dict[str, Any]:
        saved_path = self._save_uploaded_file(file)
        imported_rows, imported_sources, stock_code, years_value, sheet_name = self._load_tasks_from_excel_file(saved_path)

        parameters = [
            {
                "commission": "0.0350%",
                "xm": row[0],
                "dbbh1": row[1],
                "dbbh2": row[2],
                "zlxc": row[3],
                "zsgz": row[4],
                "ywf1": row[5],
                "ywf2": row[6],
            }
            for row in imported_rows
        ]

        return {
            "model_version": "c3",
            "sheet_name": sheet_name,
            "stock_code": stock_code,
            "recent_years": [int(years_value)] if years_value else [],
            "full_years": [],
            "parameters": parameters,
            "excel_import": {
                "original_filename": file.filename,
                "stored_file_path": str(saved_path),
                "sheet_name": sheet_name,
                "header_row": 8,
                "year_value": years_value,
                "rows": imported_sources,
            },
        }

    def _save_uploaded_file(self, file: FileStorage) -> Path:
        filename = secure_filename(file.filename or "")
        suffix = Path(filename).suffix.lower()
        if suffix not in self.ALLOWED_EXTENSIONS:
            raise ValueError("仅支持上传 .xlsx 或 .xlsm 文件")

        storage_dir = self.storage_dir()
        storage_dir.mkdir(parents=True, exist_ok=True)
        saved_path = storage_dir / f"{uuid.uuid4().hex}{suffix}"
        file.save(saved_path)
        return saved_path

    def _load_tasks_from_excel_file(self, file_path: str | Path) -> tuple[list[list[str]], list[dict[str, Any]], str, str, str]:
        file_path = str(Path(file_path).resolve())
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            workbook_values = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook["主表"] if "主表" in workbook.sheetnames else workbook.worksheets[0]
            worksheet_values = workbook_values[worksheet.title] if worksheet.title in workbook_values.sheetnames else workbook_values.worksheets[0]
        except Exception as exc:
            raise ValueError(f"Excel 文件读取失败：{exc}") from exc

        def get_merged_range(row: int, col: int):
            for merged_range in worksheet.merged_cells.ranges:
                if (
                    merged_range.min_row <= row <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col
                ):
                    return merged_range
            return None

        def normalize_label(value: Any) -> str:
            text = "" if value is None else str(value).strip()
            for token in (" ", "\n", "\r", "\t", "/", "\\", "。", "，", ",", "：", ":"):
                text = text.replace(token, "")
            return text

        def format_excel_value(value: Any) -> str:
            if value is None:
                return ""
            if isinstance(value, bool):
                return "TRUE" if value else "FALSE"
            if isinstance(value, int):
                return str(value)
            if isinstance(value, float):
                return f"{value:.15g}"
            return str(value).strip()

        allowed_binary_ops = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.Pow: operator.pow,
        }
        allowed_unary_ops = {
            ast.UAdd: operator.pos,
            ast.USub: operator.neg,
        }

        def safe_eval_expression(expr: str) -> float:
            def _eval(node):
                if isinstance(node, ast.Expression):
                    return _eval(node.body)
                if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
                    return float(node.value)
                if isinstance(node, ast.Num):
                    return float(node.n)
                if isinstance(node, ast.BinOp) and type(node.op) in allowed_binary_ops:
                    return allowed_binary_ops[type(node.op)](_eval(node.left), _eval(node.right))
                if isinstance(node, ast.UnaryOp) and type(node.op) in allowed_unary_ops:
                    return allowed_unary_ops[type(node.op)](_eval(node.operand))
                raise ValueError("不支持的公式表达式")

            parsed = ast.parse(expr, mode="eval")
            return _eval(parsed)

        def get_numeric_cell_value(row: int, col: int, visited=None) -> float:
            visited = visited or set()
            cell_key = (row, col)
            if cell_key in visited:
                raise ValueError("检测到循环公式引用")

            cached_value = worksheet_values.cell(row, col).value
            if isinstance(cached_value, (int, float)):
                return float(cached_value)

            source_value = worksheet.cell(row, col).value
            if isinstance(source_value, (int, float)):
                return float(source_value)

            if isinstance(source_value, str) and source_value.startswith("="):
                return evaluate_formula(source_value, visited | {cell_key})

            text_value = format_excel_value(cached_value if cached_value is not None else source_value)
            if not text_value:
                return 0.0
            return float(text_value)

        def evaluate_formula(formula: str, visited=None) -> float:
            expr = formula.lstrip("=").replace("^", "**")

            def replace_cell_reference(match):
                col_letters = match.group(1)
                row_number = int(match.group(2))
                col_number = column_index_from_string(col_letters)
                return str(get_numeric_cell_value(row_number, col_number, visited))

            expr = re.sub(r"\$?([A-Z]{1,3})\$?(\d+)", replace_cell_reference, expr)
            return safe_eval_expression(expr)

        def get_cell_text(row: int, col: int) -> str:
            cached_value = worksheet_values.cell(row, col).value
            source_value = worksheet.cell(row, col).value
            resolved_value = cached_value

            if resolved_value is None and isinstance(source_value, str) and source_value.startswith("="):
                try:
                    resolved_value = evaluate_formula(source_value, {(row, col)})
                except Exception:
                    resolved_value = source_value
            elif resolved_value is None:
                resolved_value = source_value

            return format_excel_value(resolved_value)

        stock_code = ""
        for row in range(1, min(worksheet.max_row, 5) + 1):
            for col in range(1, min(worksheet.max_column, 10) + 1):
                cell_value = normalize_label(worksheet.cell(row, col).value)
                if "股票" not in cell_value or "指数" not in cell_value:
                    continue

                merged_range = get_merged_range(row, col)
                stock_col = merged_range.max_col + 1 if merged_range else col + 1
                if stock_col <= worksheet.max_column:
                    stock_code = str(worksheet.cell(row, stock_col).value or "").strip()
                break
            if stock_code:
                break

        if not stock_code:
            stock_code = str(worksheet["C1"].value or "").strip()

        if not stock_code:
            raise ValueError("Excel 中没有输入股票代码")

        stock_code = stock_code.upper()

        year_header = str(worksheet["H8"].value or "").strip()
        year_digits = "".join(ch for ch in year_header if ch.isdigit())
        years_value = year_digits if year_digits in {"1", "3", "7"} else "7"

        imported_rows: list[list[str]] = []
        imported_sources: list[dict[str, Any]] = []
        current_row = 9

        while current_row <= worksheet.max_row:
            param_values: list[str] = []
            has_value = False
            for col in range(1, 8):
                text_value = get_cell_text(current_row, col)
                if text_value:
                    has_value = True
                param_values.append(text_value)

            if not has_value:
                break

            imported_rows.append(param_values)
            imported_sources.append({
                "file_path": file_path,
                "sheet_name": worksheet.title,
                "header_row": 8,
                "start_row": current_row,
                "stock_code": stock_code,
                "params": param_values.copy(),
            })
            current_row += 8

        if not imported_rows:
            raise ValueError("Excel 中未读取到参数数据，请检查模板格式")

        return imported_rows, imported_sources, stock_code, years_value, worksheet.title
