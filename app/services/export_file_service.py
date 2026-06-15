import json
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from functools import cmp_to_key
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXCEL_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# C5 分组 sheet 只展示业务关注列；kline_range 仅作为拆 sheet 的分组依据，不展示在表格里。
C5_EXPORT_COLUMNS = [
    "xm",
    "ml",
    "ReturnBeats",
    "ddBeats",
    "Return",
    "Annualized",
    "Max DD%",
    "Index Return",
    "Annualized",
    "Index max dd",
    "单位理论杠杆率收益",
    "单位实际杠杆率收益",
    "模型年化标准差",
    "指数年化标准差",
    "指数XPL",
    "模型XPL",
]

C5_EXPORT_METRIC_KEYS = ["D11", "D12", "D2", "D3", "D4", "D5", "D6", "D7", "D17", "D20"]

PERCENT_COLUMN_NAMES = {
    "ReturnBeats",
    "ddBeats",
    "Return",
    "Annualized",
    "Max DD%",
    "Index Return",
    "Index max dd",
    "单位理论杠杆率收益",
    "单位实际杠杆率收益",
}

FOUR_DECIMAL_COLUMN_NAMES = {
    "模型年化标准差",
    "指数年化标准差",
    "指数XPL",
    "模型XPL",
}

# C3 参数列名 → Google Sheet 单元格引用映射（与 _build_stock_param_result_payload 对应）
C3_PARAM_CELL_MAP = [
    ("xm",   "B6"),
    ("tp1",  "B7"),
    ("nl",   "B9"),
    ("if",   "B10"),
    ("ywfs", "B11"),
    ("ywb",  "B12"),
]
C3_PARAM_NAMES = [name for name, _ in C3_PARAM_CELL_MAP]
C3_PARAM_COUNT = len(C3_PARAM_NAMES)

# C3 指标列名 → Google Sheet 单元格引用映射（与 SQL / _build_stock_param_result_payload 对应）
C3_METRIC_CELL_MAP = [
    ("annualized_rate",      "I16"),
    ("index_annualized_rate","I19"),
    ("maxdd",                "I17"),
    ("max_index_dd",         "I20"),
    ("fee_total",            "I21"),
    ("fee_annualized",       "I22"),
    ("year_rate",            "I23"),
]

# C3 导出列定义
C3_EXPORT_COLUMNS = [
    "top choices",
    "ywf top choices",
    "data start",
    "data end",
    *C3_PARAM_NAMES,
    "annualized%",
    "index%",
    "beats",
    "strat max dd%",
    "index max dd%",
    "beats_dd",
    "Fee total",
    "Fee annualize",
    "年换手率",
    "模型年化标准差",
    "指数年化标准差",
    "指数XPL",
    "模型XPL",
]

# 导出时 "beats_dd" 在表头显示为 "beats"（与 return beats 列同名，业务约定）
C3_HEADER_DISPLAY = {"beats_dd": "beats"}

C3_COLUMN_WIDTHS = {
    "top choices": 16,
    "ywf top choices": 16,
    "data start": 14,
    "data end": 14,
    "xm": 7,
    "tp1": 7,
    "nl": 7,
    "if": 7,
    "ywfs": 7,
    "ywb": 7,
    "annualized%": 14,
    "index%": 10,
    "beats": 10,
    "beats_dd": 10,
    "strat max dd%": 14,
    "index max dd%": 14,
    "Fee total": 12,
    "Fee annualize": 14,
    "年换手率": 12,
    "模型年化标准差": 16,
    "指数年化标准差": 16,
    "指数XPL": 11,
    "模型XPL": 11,
}

C3_PERCENT_COLUMN_NAMES = {
    "annualized%",
    "index%",
    "beats",
    "beats_dd",
    "strat max dd%",
    "index max dd%",
    "Fee total",
    "Fee annualize",
    "年换手率",
}

C3_FOUR_DECIMAL_COLUMN_NAMES = {
    "模型年化标准差",
    "指数年化标准差",
    "指数XPL",
    "模型XPL",
}

C5_COLUMN_WIDTHS = {
    "xm": 7,
    "ml": 7,
    "ReturnBeats": 12,
    "ddBeats": 10,
    "Return": 10,
    "Annualized": 12,
    "Max DD%": 10,
    "Index Return": 12,
    "Index max dd": 12,
    "单位理论杠杆率收益": 18,
    "单位实际杠杆率收益": 18,
    "模型年化标准差": 16,
    "指数年化标准差": 16,
    "指数XPL": 11,
    "模型XPL": 11,
}

METRIC_DISPLAY_NAME_MAP = {
    "D2": "Return",
    "D3": "Annualized",
    "D4": "Max DD%",
    "D5": "Index Return",
    "D6": "Annualized",
    "D7": "Index max dd",
    "D8": "Fee total",
    "D9": "Fee annualized",
    "D10": "年换手率",
    "D11": "ReturnBeats",
    "D12": "ddBeats",
    "D13": "max(1y beats%)",
    "D14": "min(1y beats%)",
    "D15": "最大理论杠杆率",
    "D16": "平均理论杠杆率",
    "D17": "单位理论杠杆率收益",
    "D18": "最大实际杠杆率",
    "D19": "平均实际杠杆率",
    "D20": "单位实际杠杆率收益",
}


@dataclass(frozen=True)
class GeneratedExport:
    filename: str
    mimetype: str
    workbook: Workbook


@dataclass(frozen=True)
class WorksheetData:
    name: str
    header: list[Any]
    rows: list[list[Any]]


class TaskResultExporter(Protocol):
    """任务结果导出器协议；新增导出类型时实现这个协议并注册到 EXPORTERS。"""

    key: str

    def supports(self, task: Any) -> bool:
        ...

    def build(self, task: Any, results: list[dict[str, Any]]) -> GeneratedExport:
        ...


@dataclass(frozen=True)
class C5ResultGroup:
    """C5 原始 TaskResult 的中间结构，先归一化再展开模型行。"""

    step_index: Any
    success: bool
    stock_code: str
    kline_range: str
    xm: Any
    ml: Any
    task_id: str
    timestamp: str
    error_message: str
    models: list[dict[str, Any]]


@dataclass(frozen=True)
class C5ExportRecord:
    group: C5ResultGroup
    row: list[Any]


class C5TaskResultExporter:
    """C5 专用导出：展开每个模型为一行，并按 K 线范围拆分 worksheet。"""

    key = "google_sheet_C5"

    def supports(self, task: Any) -> bool:
        return _task_type(task) == "google_sheet_c5"

    def build(self, task: Any, results: list[dict[str, Any]]) -> GeneratedExport:
        worksheets = build_c5_worksheets(results)
        return GeneratedExport(
            filename=f"{sanitize_export_filename(_task_name(task))}.xlsx",
            mimetype=EXCEL_MIMETYPE,
            workbook=build_workbook(worksheets),
        )


class GenericTaskResultExporter:
    """通用兜底导出：未定制的任务类型也能下载原始结果。"""

    key = "generic"

    def supports(self, task: Any) -> bool:
        return True

    def build(self, task: Any, results: list[dict[str, Any]]) -> GeneratedExport:
        header = [
            "id",
            "task_id",
            "step_index",
            "success",
            "timestamp",
            "parameters",
            "result",
            "error_message",
        ]
        rows = [
            [
                item.get("id", ""),
                item.get("task_id", ""),
                item.get("step_index", ""),
                "success" if item.get("success") else "failed",
                format_time(item.get("timestamp")),
                json_cell(item.get("parameters")),
                json_cell(item.get("result")),
                item.get("error_message") or "",
            ]
            for item in results
        ]
        return GeneratedExport(
            filename=f"{sanitize_export_filename(_task_name(task))}.xlsx",
            mimetype=EXCEL_MIMETYPE,
            workbook=build_workbook([WorksheetData(name="任务结果", header=header, rows=rows)]),
        )


EXPORTERS: tuple[TaskResultExporter, ...]  # 在文件末尾初始化，确保所有导出器类已定义


def build_task_export(task: Any, results: list[dict[str, Any]]) -> GeneratedExport:
    """导出统一入口；路由层负责查任务和结果，本服务只做格式转换。"""

    exporter = get_task_result_exporter(task)
    return exporter.build(task, results)


def get_task_result_exporter(task: Any) -> TaskResultExporter:
    for exporter in EXPORTERS:
        if exporter.supports(task):
            return exporter
    task_type = getattr(task, "task_type", None) or "unknown"
    raise ValueError(f"暂不支持导出任务类型: {task_type}")


def build_c5_rows(results: list[dict[str, Any]]) -> list[list[Any]]:
    """把 C5 任务结果转换成二维表；一个参数组合下的每个模型各占一行。"""

    return [C5_EXPORT_COLUMNS] + [record.row for record in build_c5_records(results)]


def build_c5_records(results: list[dict[str, Any]]) -> list[C5ExportRecord]:
    """生成带分组上下文的导出行；上下文用于排序和拆 sheet，不写入最终 Excel。"""

    records = []
    for group in build_c5_groups(results):
        if not group.models:
            records.append(C5ExportRecord(group=group, row=c5_empty_model_row(group)))
            continue

        for model in group.models:
            records.append(C5ExportRecord(group=group, row=c5_model_row(group, model)))

    return sort_c5_records(records)


def build_c5_groups(results: list[dict[str, Any]]) -> list[C5ResultGroup]:
    """先把数据库里的松散 JSON 结果归一化，后续展开行时不用反复判空。"""

    groups = []
    for item in results:
        params = item.get("parameters") if isinstance(item.get("parameters"), dict) else {}
        result = item.get("result") if isinstance(item.get("result"), dict) else {}
        groups.append(
            C5ResultGroup(
                step_index=item.get("step_index"),
                success=bool(item.get("success")),
                stock_code=str(params.get("stock_code") or ""),
                kline_range=kline_range_from_params(params),
                xm=params.get("A1", ""),
                ml=params.get("B1", ""),
                task_id=str(item.get("task_id") or ""),
                timestamp=format_time(item.get("timestamp")),
                error_message=str(item.get("error_message") or ""),
                models=[build_c5_model(model_key, metrics) for model_key, metrics in result.items()],
            )
        )
    return groups


def build_c5_model(model_key: Any, metrics: Any) -> dict[str, Any]:
    raw_metrics = metrics if isinstance(metrics, dict) else {}
    key_text = str(model_key)
    key_parts = key_text.split("__")

    # 优先从旧键读取（向后兼容），回退到 flat_result 中读取
    start_xpl = raw_metrics.get("start_return_xpl") if isinstance(raw_metrics.get("start_return_xpl"), dict) else {}
    index_xpl = raw_metrics.get("index_return_xpl") if isinstance(raw_metrics.get("index_return_xpl"), dict) else {}

    if not start_xpl or not index_xpl:
        flat_result = raw_metrics.get("flat_result") if isinstance(raw_metrics.get("flat_result"), dict) else {}
        if not start_xpl:
            start_xpl = {
                "annual_std_dev": flat_result.get("start_annual_std_dev", ""),
                "sharpe_ratio": flat_result.get("start_sharpe_ratio", ""),
            }
        if not index_xpl:
            index_xpl = {
                "annual_std_dev": flat_result.get("index_annual_std_dev", ""),
                "sharpe_ratio": flat_result.get("index_sharpe_ratio", ""),
            }

    return {
        "model_key": key_text,
        "model_title": "__".join(key_parts[1:]) if len(key_parts) > 1 else key_text,
        "metrics": raw_metrics,
        "start_xpl": start_xpl,
        "index_xpl": index_xpl,
        "start_sharpe": start_xpl.get("sharpe_ratio", ""),
        "index_sharpe": index_xpl.get("sharpe_ratio", ""),
    }


def c5_empty_model_row(group: C5ResultGroup) -> list[Any]:
    return [
        group.xm if group.xm is not None else "",
        group.ml if group.ml is not None else "",
        *([""] * (len(C5_EXPORT_COLUMNS) - 2)),
    ]


def c5_model_row(group: C5ResultGroup, model: dict[str, Any]) -> list[Any]:
    metrics = model["metrics"]
    start_xpl = model["start_xpl"]
    index_xpl = model["index_xpl"]

    return [
        group.xm if group.xm is not None else "",
        group.ml if group.ml is not None else "",
        *[c5_metric_value(metrics, key) for key in C5_EXPORT_METRIC_KEYS],
        start_xpl.get("annual_std_dev", ""),
        index_xpl.get("annual_std_dev", ""),
        model.get("index_sharpe", ""),
        model.get("start_sharpe", ""),
    ]


def c5_metric_value(metrics: dict[str, Any], key: str) -> Any:
    """D11/D12 是导出时计算列，其它 D 指标直接读取原始结果。"""

    if key == "D11":
        return percent_diff(metrics.get("D2"), metrics.get("D5"))
    if key == "D12":
        return percent_diff(metrics.get("D4"), metrics.get("D7"))
    return metrics.get(key, "")


def c5_metric_keys() -> list[str]:
    """分组 sheet 只导出业务需要的 D 指标，顺序与 C5_EXPORT_COLUMNS 对齐。"""

    return C5_EXPORT_METRIC_KEYS[:]


def metric_sort_key(key: str) -> tuple[bool, str, int]:
    match = re.match(r"^([A-Za-z_]+)(\d+)?$", str(key))
    if not match:
        return True, str(key), 0
    prefix = match.group(1)
    number = int(match.group(2)) if match.group(2) else 0
    return prefix != "D", prefix, number


def metric_label(key: str) -> str:
    return METRIC_DISPLAY_NAME_MAP.get(str(key), str(key))


def percent_diff(left: Any, right: Any) -> str:
    left_number = to_number(left)
    right_number = to_number(right)
    if left_number is None or right_number is None:
        return ""
    return f"{left_number - right_number:.2f}%"


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(str(value).replace("%", "").strip())
    except (TypeError, ValueError):
        return None
    return None if number != number else number


def sort_c5_records(records: list[C5ExportRecord]) -> list[C5ExportRecord]:
    """按股票和 K 线区间稳定排列，sheet 内排序在分组后单独处理。"""

    def compare(left: C5ExportRecord, right: C5ExportRecord) -> int:
        comparisons = [
            compare_desc(left.group.stock_code, right.group.stock_code),
            compare_desc(parse_range_end(left.group.kline_range), parse_range_end(right.group.kline_range)),
            compare_desc(parse_range_start(left.group.kline_range), parse_range_start(right.group.kline_range)),
        ]
        return next((value for value in comparisons if value), 0)

    return sorted(records, key=cmp_to_key(compare))


def build_c5_worksheets(results: list[dict[str, Any]]) -> list[WorksheetData]:
    """按 kline_range 拆分 worksheet；同一个 K 线范围的模型行放在同一个 sheet。"""

    records = build_c5_records(results)
    if not records:
        return [WorksheetData(name="无K线区间", header=C5_EXPORT_COLUMNS, rows=[])]

    grouped_rows = group_c5_records_by_kline_range(records)

    return [
        WorksheetData(name=group_name, header=C5_EXPORT_COLUMNS, rows=group_rows)
        for group_name, group_rows in grouped_rows.items()
    ]


def group_c5_records_by_kline_range(records: list[C5ExportRecord]) -> "OrderedDict[str, list[list[Any]]]":
    grouped_records = OrderedDict()
    for record in records:
        key = str(record.group.kline_range or "无K线区间")
        grouped_records.setdefault(key, []).append(record)
    return OrderedDict(
        (key, [record.row for record in sort_c5_sheet_records(group_records)])
        for key, group_records in grouped_records.items()
    )


def sort_c5_sheet_records(records: list[C5ExportRecord]) -> list[C5ExportRecord]:
    """每个分组 sheet 内：xm 为空的置顶；两段都按 ReturnBeats 降序。"""

    return_beats_col = C5_EXPORT_COLUMNS.index("ReturnBeats")

    def compare(left: C5ExportRecord, right: C5ExportRecord) -> int:
        comparisons = [
            compare_asc(is_non_empty_xm(left.group.xm), is_non_empty_xm(right.group.xm)),
            compare_desc(parse_percent(left.row[return_beats_col]), parse_percent(right.row[return_beats_col])),
        ]
        return next((value for value in comparisons if value), 0)

    return sorted(records, key=cmp_to_key(compare))


def kline_range_from_params(params: dict[str, Any]) -> str:
    """从参数里的 kline 明细推导展示区间；缺失时统一归到 '-' 分组。"""

    kline = params.get("kline")
    if not isinstance(kline, list):
        return "-"

    dated_rows = [row for row in kline if isinstance(row, dict) and row.get("stock_date")]
    if not dated_rows:
        return "-"

    dated_rows.sort(key=lambda row: str(row.get("stock_date") or ""))
    return f"{dated_rows[0].get('stock_date', '-')} ~ {dated_rows[-1].get('stock_date', '-')}"


def build_workbook(worksheets: list[WorksheetData]) -> Workbook:
    """把通用 WorksheetData 写入 openpyxl Workbook，并统一应用表格样式。"""

    workbook = Workbook()
    workbook.remove(workbook.active)

    if not worksheets:
        worksheets = [WorksheetData(name="导出结果", header=["暂无可导出的结果"], rows=[])]

    used_names: set[str] = set()
    for index, worksheet_data in enumerate(worksheets, start=1):
        sheet = workbook.create_sheet(unique_sheet_name(worksheet_data.name, f"分组{index}", used_names))
        sheet.append([excel_cell(value) for value in worksheet_data.header])
        column_names = [str(value or "") for value in worksheet_data.header]
        for row in worksheet_data.rows:
            sheet.append([
                excel_cell(value, column_names[column_index] if column_index < len(column_names) else "")
                for column_index, value in enumerate(row)
            ])
        style_table_sheet(sheet)

    return workbook


def style_table_sheet(sheet: Any) -> None:
    """统一导出样式：冻结表头、居中显示、设置边框和自适应列宽。"""

    header_fill = PatternFill("solid", fgColor="F7E1A1")
    header_font = Font(name="Microsoft YaHei", size=10, bold=True)
    body_font = Font(name="Microsoft YaHei", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    sheet.freeze_panes = "A2"

    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 18
        for column_index in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.border = border
            if row_index == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                cell.font = body_font
                cell.alignment = body_alignment
            apply_number_format(cell, sheet.cell(row=1, column=column_index).value)

    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = column_width(sheet, column_index)


def column_width(sheet: Any, column_index: int) -> int:
    header = str(sheet.cell(row=1, column=column_index).value or "")
    if header in C5_COLUMN_WIDTHS:
        return C5_COLUMN_WIDTHS[header]
    if header in C3_COLUMN_WIDTHS:
        return C3_COLUMN_WIDTHS[header]

    max_width = 0
    for row_index in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=column_index).value
        max_width = max(max_width, display_width(value))
    return min(max(max_width + 2, 8), 40)


def display_width(value: Any) -> int:
    text = str(value or "")
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return width


def excel_cell(value: Any, column_name: str = "") -> Any:
    """openpyxl 不能写入 dict/list，复杂对象统一转成 JSON 字符串。"""

    if value is None:
        return ""
    if isinstance(value, (int, float, bool, datetime)):
        return value
    if isinstance(value, str):
        if column_name in PERCENT_COLUMN_NAMES:
            percent_value = parse_percent_cell(value)
            return "" if percent_value is None else percent_value
        numeric_value = parse_numeric_cell(value)
        return value if numeric_value is None else numeric_value
    return json_cell(value)


def apply_number_format(cell: Any, column_name: Any) -> None:
    if cell.row == 1:
        return
    column_text = str(column_name or "")
    if column_text in PERCENT_COLUMN_NAMES and isinstance(cell.value, (int, float)):
        cell.number_format = "0.00%"
    elif column_text in FOUR_DECIMAL_COLUMN_NAMES and isinstance(cell.value, (int, float)):
        cell.number_format = "0.0000"
    elif column_text in C3_PERCENT_COLUMN_NAMES and isinstance(cell.value, (int, float)):
        cell.number_format = "0.00%"
    elif column_text in C3_FOUR_DECIMAL_COLUMN_NAMES and isinstance(cell.value, (int, float)):
        cell.number_format = "0.0000"


def parse_percent_cell(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None
    if text.endswith("%"):
        number = to_number(text)
        return None if number is None else number / 100
    return parse_numeric_cell(text)


def parse_numeric_cell(value: str) -> float | int | None:
    text = value.strip()
    if not text:
        return None
    if text.endswith("%"):
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if number != number:
        return None
    return int(number) if number.is_integer() else number


def json_cell(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def format_time(value: Any) -> str:
    if not value:
        return ""
    try:
        if isinstance(value, (int, float)):
            parsed = datetime.fromtimestamp(value)
        elif isinstance(value, str):
            parsed = datetime.fromisoformat(value)
        else:
            return str(value)
    except (TypeError, ValueError):
        return str(value)
    return parsed.strftime("%Y-%m-%d %H:%M:%S")


def sanitize_export_filename(name: Any, fallback: str = "task_export") -> str:
    safe_name = "".join(char if char not in '\\/:*?"<>|' else "_" for char in str(name or "")).strip()
    return safe_name or fallback


def unique_sheet_name(label: Any, fallback: str, used_names: set[str]) -> str:
    """Excel sheet 名最多 31 个字符，且同一个 workbook 内必须唯一。"""

    base_name = sanitize_sheet_name(label, fallback)
    sheet_name = base_name
    suffix = 1
    while sheet_name in used_names:
        suffix += 1
        suffix_text = f"_{suffix}"
        sheet_name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
    used_names.add(sheet_name)
    return sheet_name


def sanitize_sheet_name(label: Any, fallback: str) -> str:
    name = str(label or "").strip() or fallback
    name = re.sub(r"[\[\]:*?/\\]", "_", name)
    name = re.sub(r"\s+", " ", name).strip("' ")
    return (name or fallback)[:31]


def parse_range_end(value: Any) -> float:
    parts = str(value or "").split("~")
    return parse_datetime_timestamp(parts[1].strip()) if len(parts) >= 2 else float("-inf")


def parse_range_start(value: Any) -> float:
    parts = str(value or "").split("~")
    return parse_datetime_timestamp(parts[0].strip()) if parts else float("-inf")


def parse_datetime_timestamp(value: str) -> float:
    try:
        return datetime.fromisoformat(value).timestamp()
    except (TypeError, ValueError):
        return float("-inf")


def parse_percent(value: Any) -> float:
    number = to_number(value)
    return number if number is not None else float("-inf")


def normalize_xm(value: Any) -> str:
    text = str(value or "").strip()
    return "" if not text or text == "0" else text


def is_non_empty_xm(value: Any) -> bool:
    return bool(normalize_xm(value))


def compare_desc(left: Any, right: Any) -> int:
    if left == right:
        return 0
    return -1 if left > right else 1


def compare_asc(left: Any, right: Any) -> int:
    if left == right:
        return 0
    return -1 if left < right else 1


def _task_type(task: Any) -> str:
    return str(getattr(task, "task_type", "") or "").lower()


# ---------------------------------------------------------------------------
# C3 导出支持
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class C3ResultGroup:
    """C3 原始 TaskResult 的中间结构。"""

    step_index: Any
    success: bool
    stock_code: str
    kline_range: str
    kline_start: str
    kline_end: str
    params: list[Any]
    task_id: str
    timestamp: str
    error_message: str
    result: dict[str, Any]


@dataclass(frozen=True)
class C3ExportRecord:
    group: C3ResultGroup
    row: list[Any]


class C3TaskResultExporter:
    """C3 专用导出：每个参数组合一行，按 K 线范围拆分 worksheet。"""

    key = "google_sheet"

    def supports(self, task: Any) -> bool:
        tt = _task_type(task)
        return tt in ("google_sheet", "google_sheet_c3")

    def build(self, task: Any, results: list[dict[str, Any]]) -> GeneratedExport:
        worksheets = build_c3_worksheets(results)
        return GeneratedExport(
            filename=f"{sanitize_export_filename(_task_name(task))}.xlsx",
            mimetype=EXCEL_MIMETYPE,
            workbook=build_workbook(worksheets),
        )


def build_c3_groups(results: list[dict[str, Any]]) -> list[C3ResultGroup]:
    groups = []
    for item in results:
        params = item.get("parameters")
        if not isinstance(params, list):
            params = []

        result = item.get("result") if isinstance(item.get("result"), dict) else {}

        # kline 数据追加在参数列表末尾（可能是 dict 或 list）
        kline_data = params[-1] if params and isinstance(params[-1], (dict, list)) else None
        numeric_params = params[:-1] if kline_data is not None else params

        kline_range = "-"
        kline_start = ""
        kline_end = ""
        if isinstance(kline_data, dict) and kline_data.get("stock_date"):
            kline_start = str(kline_data.get("stock_date", ""))
            kline_end = kline_start
            kline_range = kline_start
        elif isinstance(kline_data, list):
            dated_rows = [r for r in kline_data if isinstance(r, dict) and r.get("stock_date")]
            if dated_rows:
                dated_rows.sort(key=lambda r: str(r.get("stock_date") or ""))
                kline_start = str(dated_rows[0].get("stock_date", ""))
                kline_end = str(dated_rows[-1].get("stock_date", ""))
                kline_range = f"{kline_start} ~ {kline_end}"

        groups.append(
            C3ResultGroup(
                step_index=item.get("step_index"),
                success=bool(item.get("success")),
                stock_code=str(result.get("stock_code") or ""),
                kline_range=kline_range,
                kline_start=kline_start,
                kline_end=kline_end,
                params=numeric_params,
                task_id=str(item.get("task_id") or ""),
                timestamp=format_time(item.get("timestamp")),
                error_message=str(item.get("error_message") or ""),
                result=result,
            )
        )
    return groups


def build_c3_records(results: list[dict[str, Any]]) -> list[C3ExportRecord]:
    records = []
    for group in build_c3_groups(results):
        records.append(C3ExportRecord(group=group, row=c3_result_row(group)))
    return sort_c3_records(records)


def _c3_cell_value(result: dict[str, Any], cell_ref: str, named_key: str = "") -> Any:
    """从 C3 result dict 中读取值：优先单元格引用，回退到命名键。"""
    value = result.get(cell_ref, "")
    if value == "" and named_key:
        value = result.get(named_key, "")
    return value


def c3_result_row(group: C3ResultGroup) -> list[Any]:
    r = group.result

    # ── 参数列（从单元格引用读取，与 _build_stock_param_result_payload 一致）──
    param_values = [_c3_cell_value(r, cell, name) for name, cell in C3_PARAM_CELL_MAP]

    # ── 指标列（从单元格引用读取，与 SQL 查询字段一一对应）──
    annualized_rate      = _c3_cell_value(r, "I16", "annualized_rate")
    index_annualized_rate = _c3_cell_value(r, "I19", "index_annualized_rate")
    maxdd                = _c3_cell_value(r, "I17", "maxdd")
    max_index_dd         = _c3_cell_value(r, "I20", "max_index_dd")
    fee_total            = _c3_cell_value(r, "I21", "fee_total")
    fee_annualized       = _c3_cell_value(r, "I22", "fee_annualized")
    year_rate            = _c3_cell_value(r, "I23", "year_rate")

    # return_beats / dd_beats / turnover_rate 可能是命名键或单元格引用
    return_beats  = r.get("return_beats", "")
    dd_beats      = r.get("dd_beats", "")
    turnover_rate = r.get("turnover_rate", year_rate)  # year_rate 即年换手率

    # ── XPL 分析字段（从 flat_result 子字典读取）──
    flat = r.get("flat_result") if isinstance(r.get("flat_result"), dict) else {}
    start_monthly_std_dev = flat.get("start_monthly_std_dev", "")
    index_monthly_std_dev = flat.get("index_monthly_std_dev", "")
    index_sharpe_ratio    = flat.get("index_sharpe_ratio", "")
    start_sharpe_ratio    = flat.get("start_sharpe_ratio", "")

    # ── top choices 展示列 ──
    xm_val  = str(param_values[0]) if param_values[0] != "" else ""
    tp1_val = str(param_values[1]) if param_values[1] != "" else ""
    top_choices = f"{xm_val}/{tp1_val}" if xm_val or tp1_val else ""

    if_val   = str(param_values[3]) if param_values[3] != "" else ""
    ywfs_val = str(param_values[4]) if param_values[4] != "" else ""
    ywb_val  = str(param_values[5]) if param_values[5] != "" else ""
    ywf_parts = [v for v in (if_val, ywfs_val, ywb_val) if v]
    ywf_top_choices = "/".join(ywf_parts)

    return [
        top_choices,                      # top choices
        ywf_top_choices,                  # ywf top choices
        group.kline_start,                # data start
        group.kline_end,                  # data end
        *param_values,                    # xm, tp1, nl, if, ywfs, ywb
        annualized_rate,                  # annualized%
        index_annualized_rate,            # index%
        return_beats,                     # beats (return)
        maxdd,                            # strat max dd%
        max_index_dd,                     # index max dd%
        dd_beats,                         # beats (dd)
        fee_total,                        # Fee total
        fee_annualized,                   # Fee annualize
        turnover_rate,                    # 年换手率
        start_monthly_std_dev,            # 模型年化标准差
        index_monthly_std_dev,            # 指数年化标准差
        index_sharpe_ratio,               # 指数XPL
        start_sharpe_ratio,               # 模型XPL
    ]


def sort_c3_records(records: list[C3ExportRecord]) -> list[C3ExportRecord]:
    def compare(left: C3ExportRecord, right: C3ExportRecord) -> int:
        comparisons = [
            compare_desc(left.group.stock_code, right.group.stock_code),
            compare_desc(
                parse_range_end(left.group.kline_range),
                parse_range_end(right.group.kline_range),
            ),
        ]
        return next((value for value in comparisons if value), 0)

    return sorted(records, key=cmp_to_key(compare))


def c3_display_header(columns: list[str]) -> list[str]:
    """将内部列名转换为导出表头显示名称。"""
    return [C3_HEADER_DISPLAY.get(col, col) for col in columns]


def build_c3_worksheets(results: list[dict[str, Any]]) -> list[WorksheetData]:
    records = build_c3_records(results)
    display_header = c3_display_header(C3_EXPORT_COLUMNS)
    if not records:
        return [WorksheetData(name="无K线区间", header=display_header, rows=[])]

    grouped = OrderedDict()
    for record in records:
        key = str(record.group.kline_range or "无K线区间")
        grouped.setdefault(key, []).append(record.row)

    return [
        WorksheetData(name=group_name, header=display_header, rows=rows)
        for group_name, rows in grouped.items()
    ]


def _task_name(task: Any) -> str:
    return str(getattr(task, "name", None) or getattr(task, "id", None) or "task_export")


# 注册顺序很重要：专用导出器必须放在通用兜底导出器前面。
EXPORTERS = (
    C5TaskResultExporter(),
    C3TaskResultExporter(),
    GenericTaskResultExporter(),
)
