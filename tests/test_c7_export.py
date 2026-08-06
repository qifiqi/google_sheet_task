from types import SimpleNamespace
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from app.services.export_file_service import build_c7_stock_code_export_archive, build_task_export
from app.services.google_sheet_service_C7 import GoogleSheetService
from app.utils.c7_result_normalizer import normalize_c7_result_metrics


def test_c7_export_reads_shifted_metric_cells():
    task = SimpleNamespace(id="task-c7", name="C7 导出", task_type="google_sheet_C7")
    results = [
        {
            "id": 1,
            "task_id": "task-c7",
            "step_index": 0,
            "success": True,
            "parameters": {
                "A1": "2",
                "B1": "ml-a",
                "stock_code": "600000",
                "kline": [
                    {"stock_date": "2024-01-01"},
                    {"stock_date": "2024-12-31"},
                ],
            },
            "result": {
                "sheet-id__模型A": {
                    "D8": "10%",
                    "D9": "11%",
                    "D10": "-5%",
                    "D11": "7%",
                    "D12": "8%",
                    "D13": "-6%",
                    "D17": "4%",
                    "D18": "-1%",
                    "D21": "unused",
                    "D23": "1.23%",
                    "D26": "2.34%",
                    "flat_result": {
                        "start_monthly_std_dev": 0.1111,
                        "index_monthly_std_dev": 0.2222,
                        "index_sharpe_ratio": 1.5,
                        "start_sharpe_ratio": 2.5,
                    },
                },
            },
        }
    ]

    export = build_task_export(task, results)
    sheet = export.workbook.active
    row = [cell.value for cell in sheet[2]]

    assert row[:2] == [2, "ml-a"]
    assert row[2:13] == pytest.approx([
        0.03, 0.01, 0.03, 0.04, -0.01,
        0.1, 0.11, -0.05, 0.07, 0.08, -0.06,
    ])
    assert row[13:] == pytest.approx([
        0.0123,
        0.0234,
        0.1111,
        0.2222,
        1.5,
        2.5,
    ])


def test_c7_stock_code_export_archive_keeps_workbooks_separate():
    task = SimpleNamespace(id="task-c7", name="C7 导出", task_type="google_sheet_C7")
    results = [
        {
            "task_id": "task-c7",
            "success": True,
            "parameters": {"A1": "1", "B1": "ml-a", "stock_code": "600000", "c7_model_version": "c7_0_3"},
            "result": {"模型A": {"D2": "10%", "D3": "11%", "D4": "-5%", "D5": "7%", "D6": "8%", "D7": "-6%"}},
        },
        {
            "task_id": "task-c7",
            "success": True,
            "parameters": {"A1": "2", "B1": "ml-b", "stock_code": "600001", "c7_model_version": "c7_0_3"},
            "result": {"模型A": {"D2": "20%", "D3": "21%", "D4": "-4%", "D5": "17%", "D6": "18%", "D7": "-7%"}},
        },
    ]

    archive = build_c7_stock_code_export_archive(task, results)

    assert archive.filename == "C7 导出_按股票代码导出.zip"
    with ZipFile(archive.buffer) as zip_file:
        assert zip_file.namelist() == ["C7 导出_600001.xlsx", "C7 导出_600000.xlsx"]

        first_workbook = load_workbook(BytesIO(zip_file.read("C7 导出_600001.xlsx")), data_only=True)
        second_workbook = load_workbook(BytesIO(zip_file.read("C7 导出_600000.xlsx")), data_only=True)

    assert first_workbook.active["A2"].value == 2
    assert second_workbook.active["A2"].value == 1


def test_c5_export_uses_sheet_calculated_beats():
    task = SimpleNamespace(id="task-c5", name="C5 导出", task_type="google_sheet_C5")
    results = [{
        "task_id": "task-c5",
        "success": True,
        "parameters": {"A1": "3", "B1": "ml-a"},
        "result": {"模型A": {
            "D2": "60.62%",
            "D3": "9.95%",
            "D4": "-33.78%",
            "D5": "40.45%",
            "D6": "7.03%",
            "D7": "-32.00%",
            "D11": "2.91%",
            "D12": "-1.78%",
            "D17": "3.90%",
            "D20": "12.83%",
        }},
    }]

    row = [cell.value for cell in build_task_export(task, results).workbook.active[2]]

    assert row[2:5] == pytest.approx([0.2017, -0.0178, 0.0292])
    assert row[5:7] == pytest.approx([0.0291, -0.0178])


def test_c7_export_preserves_shifted_beats_and_unformatted_percentages():
    task = SimpleNamespace(id="task-c7", name="C7 导出", task_type="google_sheet_C7")
    results = [{
        "task_id": "task-c7",
        "success": True,
        "parameters": {"A1": "4", "B1": "ml-a"},
        "result": {"模型A": {
            "D8": "-3.39%",
            "D9": "-1.14%",
            "D10": "-0.78",
            "D11": "192.05%",
            "D12": "42.89%",
            "D13": "-47.78%",
            "D17": "-44.04%",
            "D18": "-0.30",
            "D23": "-0.32%",
            "D26": "-1.02%",
        }},
    }]

    row = [cell.value for cell in build_task_export(task, results).workbook.active[2]]

    assert row[2:15] == pytest.approx([
        -1.9544, -0.3022, -0.44, -0.4404, -0.3,
        -0.0339, -0.0114, -0.78, 1.9205, 0.4289, -0.4778,
        -0.0032, -0.0102,
    ])


def test_c7_result_payload_reads_d8_to_d26_and_normalizes_percentages():
    service = object.__new__(GoogleSheetService)
    service.task_id = "task-c7"
    result = {
        "模型A": {
            "D8": "-3.39%", "D9": "-1.14%", "D10": "-0.78",
            "D11": "192.05%", "D12": "42.89%", "D13": "-47.78%",
            "D14": "0.67%", "D15": "0.00", "D16": "55.49",
            "D17": "-44.04%", "D18": "-0.30", "D19": "0.21",
            "D20": "-71.95%", "D21": "4.00", "D22": "358.65%",
            "D23": "-0.32%", "D24": "400.29%", "D25": "112.51%", "D26": "-1.02%",
        }
    }

    payload = service._build_stock_param_result_payload(
        "C7 测试", 0, {"A1": "4", "B1": "ml-a", "kline": []}, result,
    )

    assert payload["return_rate"] == pytest.approx(-0.0339)
    assert payload["maxdd"] == pytest.approx(-0.78)
    assert payload["return_beats"] == pytest.approx(-0.4404)
    assert payload["dd_beats"] == pytest.approx(-0.3)
    assert payload["fee_annualized"] == pytest.approx(0)
    assert payload["max_1y_beats"] == pytest.approx(0.21)
    assert payload["max_theoretical_leverage"] == "4.00"
    assert payload["avg_theoretical_leverage"] == pytest.approx(3.5865)
    assert payload["max_actual_leverage"] == pytest.approx(4.0029)
    assert payload["avg_actual_leverage"] == pytest.approx(1.1251)


def test_normalize_c7_result_metrics_matches_c5_units():
    result = normalize_c7_result_metrics({
        "D10": "-0.88",
        "D15": "0.00",
        "D18": "-0.06",
        "D19": "0.48",
        "D21": "3.50",
        "D22": "295.34%",
        "D24": "350.24%",
        "D25": "93.62%",
    })

    assert result["D10"] == "-88.00%"
    assert result["D15"] == "0.00%"
    assert result["D18"] == "-6.00%"
    assert result["D19"] == "48.00%"
    assert result["D21"] == "3.50"
    assert result["D22"] == pytest.approx(2.9534)
    assert result["D24"] == pytest.approx(3.5024)
    assert result["D25"] == pytest.approx(0.9362)
