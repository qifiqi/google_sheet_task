from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


SCRIPT_PATH = (
    Path(__file__).resolve().parents[1]
    / "scripts"
    / "download_merge_backtest_multi_product_previews.py"
)


def load_script_module():
    spec = importlib.util.spec_from_file_location("multi_product_export_merge", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def save_workbook(path: Path, marker: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "多品全局预览"
    sheet["A1"] = marker
    sheet["A1"].font = Font(bold=True)
    sheet["B2"] = 123
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "merged"
    workbook.save(path)
    workbook.close()


def test_discover_groups_pairs_files_by_module_and_period(tmp_path):
    module = load_script_module()
    save_workbook(tmp_path / "alpha_1y.xlsx", "alpha 1y")
    save_workbook(tmp_path / "alpha_3y.xlsx", "alpha 3y")
    save_workbook(tmp_path / "ignored.xlsx", "ignored")

    groups, warnings = module.discover_workbook_groups(tmp_path)

    assert [group.module_key for group in groups] == ["alpha"]
    assert groups[0].files_by_period["1y"].name == "alpha_1y.xlsx"
    assert groups[0].files_by_period["3y"].name == "alpha_3y.xlsx"
    assert any("未唯一匹配" in warning for warning in warnings)


def test_merge_directory_creates_1y_and_3y_sheets(tmp_path):
    module = load_script_module()
    save_workbook(tmp_path / "alpha_1y.xlsx", "alpha 1y")
    save_workbook(tmp_path / "alpha_3y.xlsx", "alpha 3y")

    generated = module.merge_directory(tmp_path, tmp_path / "merged", overwrite=True)

    assert len(generated) == 1
    assert generated[0].name == "alpha.xlsx"
    merged_workbook = load_workbook(generated[0])
    try:
        assert merged_workbook.sheetnames == ["1y", "3y"]
        assert merged_workbook["1y"]["A1"].value == "alpha 1y"
        assert merged_workbook["3y"]["A1"].value == "alpha 3y"
        assert merged_workbook["1y"]["A1"].font.bold is True
        assert "A3:B3" in [str(item) for item in merged_workbook["1y"].merged_cells.ranges]
    finally:
        merged_workbook.close()
