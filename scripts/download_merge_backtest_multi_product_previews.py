"""Download and merge multi-product backtest preview Excel files.

默认指向远程站点 http://172.18.20.17:5001。

脚本会批量导出多品数据回测的全局预览 Excel，然后按文件名中的
``1y`` / ``3y`` 配对合并：同一模块生成一个工作簿，sheet 名分别为
``1y`` 和 ``3y``。
"""

from __future__ import annotations

import argparse
import copy
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


DEFAULT_BASE_URL = "http://172.18.20.17:5001"
DEFAULT_OUTPUT_DIR = "../downloads/backtest_multi_product_previews"
TASK_TYPE = "backtest_multi_product"
XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PERIODS = ("1y", "3y")


@dataclass(frozen=True)
class BacktestTask:
    id: str
    name: str
    status: str
    created_at: str


@dataclass(frozen=True)
class DownloadResult:
    task_id: str
    ok: bool
    path: Path | None = None
    error: str = ""


@dataclass(frozen=True)
class WorkbookGroup:
    module_key: str
    files_by_period: dict[str, Path]


def normalize_base_url(raw_url: str) -> str:
    value = (raw_url or DEFAULT_BASE_URL).strip().rstrip("/")
    parsed = urlparse(value)
    if not parsed.scheme or not parsed.netloc:
        raise ValueError(f"无效 base-url: {raw_url}")
    return f"{parsed.scheme}://{parsed.netloc}"


def build_session(timeout: float, retries: int) -> requests.Session:
    retry = Retry(
        total=retries,
        connect=retries,
        read=retries,
        status=retries,
        backoff_factor=0.8,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "POST"}),
    )
    adapter = HTTPAdapter(max_retries=retry)
    session = requests.Session()
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update({"User-Agent": "backtest-multi-product-exporter/1.0"})
    session.request_timeout = timeout  # type: ignore[attr-defined]
    return session


def request_json(session: requests.Session, method: str, url: str, **kwargs: Any) -> dict[str, Any]:
    timeout = getattr(session, "request_timeout", 30.0)
    response = session.request(method, url, timeout=timeout, **kwargs)
    if response.status_code == 401:
        raise RuntimeError("认证失败：请提供 --username/--password 或 --access-token")
    if response.status_code == 403:
        raise RuntimeError(f"权限不足：{response.text[:300]}")
    response.raise_for_status()
    try:
        return response.json()
    except json.JSONDecodeError as err:
        raise RuntimeError(f"接口未返回 JSON: {url}") from err


def login(session: requests.Session, base_url: str, username: str, password: str) -> str:
    payload = request_json(
        session,
        "POST",
        urljoin(base_url, "/api/auth/login"),
        json={"username": username, "password": password},
    )
    if payload.get("code") != 0:
        raise RuntimeError(payload.get("message") or "登录失败")
    token = str(((payload.get("data") or {}).get("access_token") or "")).strip()
    if not token:
        raise RuntimeError("登录成功响应中没有 access_token")
    return token


def configure_auth(session: requests.Session, base_url: str, args: argparse.Namespace) -> None:
    token = (args.access_token or os.environ.get("BACKTEST_ACCESS_TOKEN") or "").strip()
    username = (args.username or os.environ.get("BACKTEST_USERNAME") or "").strip()
    password = args.password or os.environ.get("BACKTEST_PASSWORD") or ""

    if not token and username and password:
        token = login(session, base_url, username, password)

    if token:
        session.headers.update({"Authorization": f"Bearer {token}"})


def parse_datetime(value: str) -> datetime | None:
    raw = (value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed
    return parsed.astimezone().replace(tzinfo=None)


def list_multi_product_tasks(
    session: requests.Session,
    base_url: str,
    per_page: int,
    status: str | None,
    keyword: str,
    max_pages: int | None,
    created_after: datetime | None,
) -> list[BacktestTask]:
    tasks: list[BacktestTask] = []
    page = 1

    while True:
        params: dict[str, Any] = {
            "task_type": TASK_TYPE,
            "page": page,
            "per_page": per_page,
        }
        if status:
            params["status"] = status
        if keyword:
            params["keyword"] = keyword

        payload = request_json(session, "GET", urljoin(base_url, "/api/tasks"), params=params)
        should_stop = False
        for item in payload.get("tasks") or []:
            task_id = str(item.get("id") or "").strip()
            if not task_id:
                continue
            created_at = str(item.get("created_at") or "")
            created_time = parse_datetime(created_at)
            if created_after is not None and created_time is not None and created_time < created_after:
                should_stop = True
                continue
            tasks.append(
                BacktestTask(
                    id=task_id,
                    name=str(item.get("name") or task_id),
                    status=str(item.get("status") or ""),
                    created_at=created_at,
                )
            )

        pagination = payload.get("pagination") or {}
        if should_stop or not bool(pagination.get("has_next")):
            break
        if max_pages is not None and page >= max_pages:
            break
        page += 1

    return tasks


def parse_filename_from_content_disposition(value: str) -> str:
    if not value:
        return ""

    filename_star = re.search(r"filename\*=UTF-8''([^;]+)", value, flags=re.IGNORECASE)
    if filename_star:
        return unquote(filename_star.group(1).strip().strip('"'))

    filename = re.search(r'filename="?([^";]+)"?', value, flags=re.IGNORECASE)
    if filename:
        return filename.group(1).strip()

    return ""


def safe_filename(name: str, fallback: str) -> str:
    raw = (name or fallback).strip() or fallback
    cleaned = re.sub(r'[\\/:*?"<>|\r\n\t]+', "_", raw).strip(" .")
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned[:180] or fallback


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 10_000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成不重复文件名: {path}")


def download_global_preview(
    session: requests.Session,
    base_url: str,
    task: BacktestTask,
    output_dir: Path,
    overwrite: bool,
) -> DownloadResult:
    url = urljoin(base_url, f"/backtest-multi-product/api/global-preview/{task.id}/export")
    timeout = getattr(session, "request_timeout", 30.0)

    try:
        response = session.get(url, timeout=timeout)
        if response.status_code == 401:
            return DownloadResult(task_id=task.id, ok=False, error="认证失败")
        if response.status_code == 403:
            return DownloadResult(task_id=task.id, ok=False, error="权限不足")
        response.raise_for_status()

        content_type = response.headers.get("Content-Type", "")
        if XLSX_MIMETYPE not in content_type and not response.content.startswith(b"PK"):
            return DownloadResult(
                task_id=task.id,
                ok=False,
                error=f"响应不是 xlsx: content-type={content_type}, body={response.text[:200]}",
            )

        header_filename = parse_filename_from_content_disposition(
            response.headers.get("Content-Disposition", "")
        )
        filename = safe_filename(header_filename or f"{task.name or task.id}.xlsx", f"{task.id}.xlsx")
        if not filename.lower().endswith(".xlsx"):
            filename = f"{filename}.xlsx"

        path = output_dir / filename
        if not overwrite:
            path = unique_path(path)
        path.write_bytes(response.content)
        return DownloadResult(task_id=task.id, ok=True, path=path)
    except requests.RequestException as err:
        return DownloadResult(task_id=task.id, ok=False, error=str(err))


def detect_period(path: Path) -> str | None:
    stem = path.stem.lower()
    found = [period for period in PERIODS if period in stem]
    if len(found) == 1:
        return found[0]
    return None


def build_module_key(path: Path, period: str) -> str:
    stem = path.stem
    key = re.sub(re.escape(period), "", stem, count=1, flags=re.IGNORECASE)
    key = re.sub(r"[_\-\s()\[\]{}]+", "_", key)
    key = key.strip("_. -")
    return key or stem


def discover_workbook_groups(input_dir: Path) -> tuple[list[WorkbookGroup], list[str]]:
    buckets: dict[str, dict[str, list[Path]]] = {}
    warnings: list[str] = []

    for path in sorted(input_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        period = detect_period(path)
        if period is None:
            warnings.append(f"跳过，文件名未唯一匹配 1y/3y: {path.name}")
            continue
        module_key = build_module_key(path, period)
        buckets.setdefault(module_key, {}).setdefault(period, []).append(path)

    groups: list[WorkbookGroup] = []
    for module_key, files_by_period in sorted(buckets.items()):
        missing = [period for period in PERIODS if period not in files_by_period]
        if missing:
            warnings.append(f"跳过 {module_key}，缺少: {', '.join(missing)}")
            continue

        duplicate_periods = {
            period: files for period, files in files_by_period.items() if len(files) > 1
        }
        if duplicate_periods:
            details = ", ".join(
                f"{period}={len(files)} 个文件" for period, files in duplicate_periods.items()
            )
            warnings.append(f"跳过 {module_key}，同一周期存在多个文件: {details}")
            continue

        groups.append(
            WorkbookGroup(
                module_key=module_key,
                files_by_period={period: files_by_period[period][0] for period in PERIODS},
            )
        )

    return groups, warnings


def copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy.copy(source_cell.protection)


def copy_sheet(source_sheet: Worksheet, target_sheet: Worksheet) -> None:
    for row in source_sheet.iter_rows():
        for source_cell in row:
            copy_cell(source_cell, target_sheet.cell(source_cell.row, source_cell.column))

    for row_index, dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_index].height = dimension.height
        target_sheet.row_dimensions[row_index].hidden = dimension.hidden

    for column_letter, dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[column_letter]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden

    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    if source_sheet.auto_filter.ref:
        target_sheet.auto_filter.ref = source_sheet.auto_filter.ref


def copy_first_sheet_to_workbook(source_path: Path, target_workbook: Workbook, sheet_name: str) -> None:
    source_workbook = load_workbook(source_path)
    try:
        source_sheet = source_workbook.worksheets[0]
        target_sheet = target_workbook.create_sheet(sheet_name)
        copy_sheet(source_sheet, target_sheet)
    finally:
        source_workbook.close()


def merge_group(group: WorkbookGroup, output_dir: Path, overwrite: bool) -> Path:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    try:
        for period in PERIODS:
            copy_first_sheet_to_workbook(group.files_by_period[period], workbook, period)

        output_dir.mkdir(parents=True, exist_ok=True)
        output_name = safe_filename(f"{group.module_key}.xlsx", "merged.xlsx")
        output_path = output_dir / output_name
        if not overwrite:
            output_path = unique_path(output_path)
        workbook.save(output_path)
        return output_path
    finally:
        workbook.close()


def merge_directory(input_dir: Path, output_dir: Path, overwrite: bool) -> list[Path]:
    if not input_dir.exists():
        raise FileNotFoundError(f"目录不存在: {input_dir}")
    if not input_dir.is_dir():
        raise NotADirectoryError(f"不是目录: {input_dir}")

    groups, warnings = discover_workbook_groups(input_dir)
    for warning in warnings:
        print(f"WARN: {warning}", file=sys.stderr)

    if not groups:
        raise FileNotFoundError(f"目录下没有找到可按 1y/3y 配对合并的 Excel: {input_dir}")

    generated: list[Path] = []
    for group in groups:
        output_path = merge_group(group, output_dir, overwrite=overwrite)
        generated.append(output_path)
        print(f"已合并: {group.files_by_period['1y'].name} + {group.files_by_period['3y'].name} -> {output_path}")

    return generated


def write_manifest(
    output_dir: Path,
    tasks: list[BacktestTask],
    results: list[DownloadResult],
    merged_files: list[Path],
) -> None:
    by_task_id = {result.task_id: result for result in results}
    rows = []
    for task in tasks:
        result = by_task_id.get(task.id)
        rows.append(
            {
                "task_id": task.id,
                "task_name": task.name,
                "task_status": task.status,
                "created_at": task.created_at,
                "downloaded": bool(result and result.ok),
                "path": str(result.path) if result and result.path else "",
                "error": result.error if result else "未处理",
            }
        )

    payload = {
        "downloads": rows,
        "merged_files": [str(path) for path in merged_files],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="批量导出多品数据回测 Excel，并把同模块 1y/3y 文件合并成双 sheet 工作簿。",
    )
    parser.add_argument(
        "--base-url",
        default=os.environ.get("BACKTEST_BASE_URL", DEFAULT_BASE_URL),
        help=f"站点地址，默认: {DEFAULT_BASE_URL}",
    )
    parser.add_argument("--username", help="登录用户名；也可用 BACKTEST_USERNAME")
    parser.add_argument("--password", help="登录密码；也可用 BACKTEST_PASSWORD")
    parser.add_argument("--access-token", help="浏览器 localStorage 里的 access_token；也可用 BACKTEST_ACCESS_TOKEN")
    parser.add_argument(
        "--output-dir",
        default=DEFAULT_OUTPUT_DIR,
        help=f"下载根目录，默认: {DEFAULT_OUTPUT_DIR}/<时间戳>",
    )
    parser.add_argument(
        "--merge-only",
        type=Path,
        help="只合并已有下载目录，不访问远程接口。",
    )
    parser.add_argument("--skip-merge", action="store_true", help="只下载，不执行 1y/3y 合并")
    parser.add_argument("--per-page", type=int, default=20, help="任务分页大小，默认 20")
    parser.add_argument("--status", default="completed", help="只下载指定状态任务，默认 completed")
    parser.add_argument("--all-statuses", action="store_true", help="不过滤任务状态")
    parser.add_argument("--keyword", default="", help="按任务列表 keyword 过滤")
    parser.add_argument("--days", type=int, default=0, help="只下载最近 N 天创建的任务，默认 0 表示不过滤")
    parser.add_argument("--max-pages", type=int, help="最多扫描多少页，调试用")
    parser.add_argument("--timeout", type=float, default=60.0, help="单次请求超时秒数，默认 60")
    parser.add_argument("--retries", type=int, default=3, help="网络重试次数，默认 3")
    parser.add_argument("--overwrite", action="store_true", help="同名文件存在时覆盖")
    parser.add_argument("--dry-run", action="store_true", help="只列出任务，不下载")
    return parser.parse_args()


def run_merge_only(args: argparse.Namespace) -> int:
    input_dir = args.merge_only.resolve()
    output_dir = input_dir / "merged"
    generated = merge_directory(input_dir, output_dir, overwrite=args.overwrite)
    print(f"合并完成: {len(generated)} 个文件")
    print(f"合并目录: {output_dir.resolve()}")
    return 0


def main() -> int:
    args = parse_args()
    if args.merge_only:
        return run_merge_only(args)

    base_url = normalize_base_url(args.base_url)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(args.output_dir) / timestamp
    export_dir = output_dir / "exports"
    merge_dir = output_dir / "merged"
    export_dir.mkdir(parents=True, exist_ok=True)

    session = build_session(timeout=args.timeout, retries=args.retries)
    configure_auth(session, base_url, args)

    started_at = time.monotonic()
    print(f"站点: {base_url}")
    print(f"下载目录: {export_dir.resolve()}")

    created_after = None
    if args.days and args.days > 0:
        created_after = datetime.now() - timedelta(days=args.days)
        print(f"时间过滤: 仅下载 {created_after:%Y-%m-%d %H:%M:%S} 之后创建的任务")

    tasks = list_multi_product_tasks(
        session=session,
        base_url=base_url,
        per_page=max(1, min(args.per_page, 500)),
        status=None if args.all_statuses else args.status,
        keyword=args.keyword.strip(),
        max_pages=args.max_pages,
        created_after=created_after,
    )
    print(f"找到 {len(tasks)} 个多品回测任务")

    if args.dry_run:
        for task in tasks:
            print(f"- {task.id} | {task.status} | {task.name}")
        return 0

    results: list[DownloadResult] = []
    for index, task in enumerate(tasks, start=1):
        print(f"[{index}/{len(tasks)}] 下载 {task.id} {task.name}")
        result = download_global_preview(
            session=session,
            base_url=base_url,
            task=task,
            output_dir=export_dir,
            overwrite=args.overwrite,
        )
        results.append(result)
        if result.ok:
            print(f"  OK: {result.path}")
        else:
            print(f"  FAIL: {result.error}", file=sys.stderr)

    merged_files: list[Path] = []
    if not args.skip_merge:
        try:
            merged_files = merge_directory(export_dir, merge_dir, overwrite=args.overwrite)
        except FileNotFoundError as err:
            print(f"WARN: {err}", file=sys.stderr)

    write_manifest(output_dir, tasks, results, merged_files)
    ok_count = sum(1 for result in results if result.ok)
    fail_count = len(results) - ok_count
    elapsed = time.monotonic() - started_at
    print(f"完成: 下载成功 {ok_count}, 下载失败 {fail_count}, 合并 {len(merged_files)}, 用时 {elapsed:.1f}s")
    print(f"清单: {(output_dir / 'manifest.json').resolve()}")
    return 0 if fail_count == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
