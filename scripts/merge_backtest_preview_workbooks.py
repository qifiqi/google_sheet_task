"""Merge a rerun backtest preview workbook into a main preview workbook."""

from __future__ import annotations

import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_RERUN_SUMMARY_LABEL = "近5年重跑"
DEFAULT_SUMMARY_SUFFIX = ""
DEFAULT_INPUT_DIR = (
    Path(__file__).resolve().parents[1]
    / "downloads"
    / "backtest_global_previews"
    / "20260525_100439"
)
OVERWRITE_OUTPUT = True


def copy_cell(source_cell, target_cell) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy.copy(source_cell.protection)


def copy_row(source_sheet: Worksheet, target_sheet: Worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, source_sheet.max_column + 1):
        copy_cell(source_sheet.cell(source_row, column), target_sheet.cell(target_row, column))
    if source_row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height


def copy_sheet(source_sheet: Worksheet, target_sheet: Worksheet) -> None:
    for row in range(1, source_sheet.max_row + 1):
        copy_row(source_sheet, target_sheet, row, row)

    for column_letter, dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[column_letter]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden

    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines


def append_summary_rows(
    source_sheet: Worksheet,
    target_sheet: Worksheet,
    summary_label: str,
) -> None:
    if source_sheet.max_row < 2:
        return

    start_row = target_sheet.max_row + 1
    for offset, source_row in enumerate(range(2, source_sheet.max_row + 1)):
        target_row = start_row + offset
        copy_row(source_sheet, target_sheet, source_row, target_row)
        if offset == 0:
            original_period = target_sheet.cell(target_row, 1).value
            if summary_label:
                target_sheet.cell(target_row, 1).value = (
                    f"{original_period}（{summary_label}）" if original_period else summary_label
                )
            else:
                target_sheet.cell(target_row, 1).value = original_period

    row_offset = start_row - 2
    for merged_range in source_sheet.merged_cells.ranges:
        source_range = CellRange(str(merged_range))
        if source_range.min_row < 2:
            continue
        target_sheet.merge_cells(
            start_row=source_range.min_row + row_offset,
            start_column=source_range.min_col,
            end_row=source_range.max_row + row_offset,
            end_column=source_range.max_col,
        )


def unique_sheet_name(workbook, desired_name: str) -> str:
    if desired_name not in workbook.sheetnames:
        return desired_name

    for index in range(2, 1000):
        candidate = f"{desired_name}_{index}"
        if candidate not in workbook.sheetnames:
            return candidate

    raise RuntimeError(f"无法生成不重复 sheet 名: {desired_name}")


def merge_workbooks(
    main_path: Path,
    rerun_path: Path,
    output_path: Path,
    summary_label: str,
    overwrite: bool,
) -> None:
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"输出文件已存在: {output_path}")

    main_workbook = load_workbook(main_path)
    rerun_workbook = load_workbook(rerun_path)

    try:
        if "汇总" not in main_workbook.sheetnames:
            raise ValueError(f"主文件缺少 汇总 sheet: {main_path}")
        if "汇总" not in rerun_workbook.sheetnames:
            raise ValueError(f"近5年文件缺少 汇总 sheet: {rerun_path}")

        append_summary_rows(
            source_sheet=rerun_workbook["汇总"],
            target_sheet=main_workbook["汇总"],
            summary_label=summary_label,
        )

        for sheet_name in rerun_workbook.sheetnames:
            if sheet_name == "汇总":
                continue
            copied_sheet = main_workbook.create_sheet(unique_sheet_name(main_workbook, sheet_name))
            copy_sheet(rerun_workbook[sheet_name], copied_sheet)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        main_workbook.save(output_path)
    finally:
        main_workbook.close()
        rerun_workbook.close()


def build_rerun_main_pairs(input_dir: Path) -> list[tuple[Path, Path]]:
    rerun_files = sorted(input_dir.glob(f"*({DEFAULT_RERUN_SUMMARY_LABEL}).xlsx"))
    pairs: list[tuple[Path, Path]] = []

    for rerun_file in rerun_files:
        main_stem = rerun_file.stem.replace(f" ({DEFAULT_RERUN_SUMMARY_LABEL})", "")
        main_file = rerun_file.with_name(f"{main_stem}.xlsx")
        if main_file.exists():
            pairs.append((main_file, rerun_file))
        else:
            print(f"跳过，找不到主文件: {main_file}")

    return pairs


def merge_directory(
    input_dir: Path = DEFAULT_INPUT_DIR,
    summary_label: str = DEFAULT_SUMMARY_SUFFIX,
    overwrite: bool = OVERWRITE_OUTPUT,
) -> list[Path]:
    input_dir = input_dir.resolve()
    if not input_dir.exists():
        raise FileNotFoundError(f"目录不存在: {input_dir}")
    if not input_dir.is_dir():
        raise NotADirectoryError(f"不是目录: {input_dir}")

    generated_files: list[Path] = []
    pairs = build_rerun_main_pairs(input_dir)
    if not pairs:
        raise FileNotFoundError(f"目录下没有找到可合并的近5年重跑文件: {input_dir}")

    for main_file, rerun_file in pairs:
        output = main_file.with_name(f"{main_file.stem}_merged{main_file.suffix}")
        merge_workbooks(
            main_path=main_file,
            rerun_path=rerun_file,
            output_path=output,
            summary_label=summary_label,
            overwrite=overwrite,
        )
        generated_files.append(output)
        print(f"已生成: {output}")

    return generated_files


def main() -> int:
    merge_directory()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
